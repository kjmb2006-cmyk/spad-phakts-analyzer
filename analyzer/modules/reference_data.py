"""
SPAD Analyzer — Référentiel organisationnel et cibles par formulaire

Sources (fournies par l'utilisateur, phase pilote 2026), stockées dans
data/reference/ :
  - org_unit.xlsx : référentiel région / district / établissement / enquêteur /
    superviseur — dérivé de « fichiers_org_unit_vf.xlsx » (liste "choices" du
    XLSForm, celle réellement proposée aux enquêteurs sur le terrain).
    ⚠️ Version EXPURGÉE des noms d'enquêteurs/superviseurs (remplacés par leur
    seul code, ex. D01ENQ1) avant tout commit — le dépôt GitHub est public et
    ces noms sont des données personnelles. Les établissements gardent leur
    nom complet (ce ne sont pas des données personnelles). Le fichier source
    original (avec noms) reste uniquement sur le poste local de l'utilisateur.
  - noms_personnel.local.json (optionnel, non versionné — voir .gitignore) :
    complète les noms réels des enquêteurs/superviseurs pour l'affichage LOCAL
    uniquement (vues « par enquêteur »/« par superviseur »). Absent par
    défaut (dépôt public, machines tierces, build packagée) → l'app retombe
    alors sur le code (comportement actuel, sans régression de confidentialité).
  - tirage_etablissements.xlsx : dérivé de
    « tirage_10_etsa_par_district_kd_vu.xlsx » — tirage au sort des 120
    établissements (méthodologie : 1 EPHR/EPHD + 4 CSU + 5 CSR par district,
    seed=20260713), avec le comptage SIG des décès maternels par
    établissement. Ne contient aucune donnée personnelle.

Règles de cible par formulaire — déduites empiriquement de l'observation de
spadapp-zeta.vercel.app (aucun document méthodologique séparé fourni) et
vérifiées ici par recoupement avec les totaux nationaux déjà observés :
  F5  (Tabac — Femmes)       : 15 par établissement, fixe   → total national 1800 (exact)
  F6  (Tabac — Personnel)    : 3 si EPH/CSU, 2 si CSR-DM, 1 si CSR-D
                                → total calculé 271 (≈ 273 observé — écart de
                                mémorisation probable, à confirmer)
  F7  (Vaccination Ménages)  : 15 par établissement, fixe   → total national 1800 (exact)
  F8  (Vaccination Étab.)    : 1 par établissement, fixe    → total national 120 (exact)
  F01 (RDM — District)       : 1 par district, fixe         → total national 12 (exact)
  F02 (RDM — Établissement)  : 1 fiche par établissement ayant enregistré AU
                                MOINS un décès maternel au SIG (pas 1 par
                                décès — EPHR SAN-PEDRO a 43 décès SIG mais
                                cible F02 = 1 pour cet établissement, comme
                                observé sur spadapp-zeta.vercel.app), et PAS
                                1 par établissement de l'échantillon : les
                                établissements sans décès SIG n'ont rien à
                                remplir → total national = nombre
                                d'établissements avec sig_deces_maternels > 0
                                (16 avec tirage_etablissements.xlsx en l'état
                                — à recouper avec le DHIS2 si un écart
                                subsiste, ex. 17 attendu sur le terrain)
  F07 (RDM — Grille)         : somme des décès maternels notifiés au SIG pour
                                les établissements du district — c'est un
                                plancher, le dépasser est normal → total
                                calculé 187 (≈ 188 observé)

Ces règles sont marquées « à confirmer » dans l'UI tant qu'aucun document
officiel ne les a validées explicitement — elles reproduisent fidèlement les
totaux déjà observés, mais restent une reconstruction, pas une source
primaire.
"""
import os
import re
import json
import openpyxl

from modules import forms_registry

_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
_ANALYZER_DIR = os.path.dirname(_MODULE_DIR)
DEFAULT_ORG_UNIT_PATH = os.path.join(_ANALYZER_DIR, 'data', 'reference', 'org_unit.xlsx')
DEFAULT_TIRAGE_PATH   = os.path.join(_ANALYZER_DIR, 'data', 'reference', 'tirage_etablissements.xlsx')
DEFAULT_LOCAL_NAMES_PATH = os.path.join(_ANALYZER_DIR, 'data', 'reference', 'noms_personnel.local.json')
RENDER_SECRET_NAMES_PATH = '/etc/secrets/noms_personnel.local.json'

# FORM_CODES / FORM_LABELS / FORM_NAME_HINT ne sont plus des constantes figées
# — elles reflètent en direct le registre modifiable (modules/forms_registry.py,
# écran /admin/forms) : formulaires SPAD historiques activables/désactivables,
# et formulaires supplémentaires ajoutés sans toucher au code. __getattr__
# (PEP 562) permet à `rd.FORM_CODES` etc. de continuer à se lire comme avant
# partout dans le code, sans modifier chaque site d'appel.
def __getattr__(name):
    if name == 'FORM_CODES':
        return forms_registry.active_codes()
    if name == 'FORM_LABELS':
        return forms_registry.labels()
    if name == 'FORM_NAME_HINT':
        return forms_registry.name_hints()
    raise AttributeError(f"module {__name__!r} has no attribute {name!r}")


_cache = {}  # (org_unit_path, tirage_path) -> référentiel chargé


def guess_form_type(name):
    """Devine le code SPAD correspondant à un nom de formulaire Kobo, à
    partir de la même convention de nommage que FORM_NAME_HINT (ex.
    « 5_PNLTA_SPAD_Fiche_Femmes... » → F5). Utilisé pour pré-sélectionner
    le type dans la page « Suivi multi-formulaires » plutôt que de partir
    d'un menu vide entièrement à la main — source d'erreurs silencieuses
    (voir validate_mapping). Renvoie None si aucun préfixe connu n'est
    trouvé (formulaire non-SPAD, ou nommé différemment)."""
    if not name:
        return None
    name_low = name.lower()
    for code, hint in forms_registry.name_hints().items():
        if hint.lower() in name_low:
            return code
    return None


def validate_mapping(mapping, assets):
    """Détecte les correspondances SPAD ↔ Kobo suspectes avant qu'elles ne
    faussent silencieusement les taux de complétude — un utilisateur peut
    se tromper de formulaire dans un menu déroulant sans s'en apercevoir.

    Renvoie (erreurs, avertissements) :
      - erreurs : le même formulaire Kobo utilisé pour 2 codes SPAD
        différents — jamais légitime (chaque formulaire SPAD doit
        correspondre à un formulaire Kobo distinct), à corriger avant tout
        calcul.
      - avertissements : le nom du formulaire sélectionné ne contient pas
        le préfixe attendu pour ce code (voir FORM_NAME_HINT) — heuristique
        fondée sur la convention de nommage observée, pas une certitude
        (un formulaire nommé différemment resterait valide) : signalé,
        mais ne bloque pas le calcul."""
    by_uid = {a['uid']: a.get('name', '') for a in assets}

    vus = {}
    for code, uid in mapping.items():
        vus.setdefault(uid, []).append(code)
    erreurs = [
        f"{' et '.join(codes)} pointent vers le même formulaire Kobo "
        f"(« {by_uid.get(uid, uid)} ») — corrigez avant de calculer."
        for uid, codes in vus.items() if len(codes) > 1
    ]

    hints = forms_registry.name_hints()
    labels = forms_registry.labels()
    avertissements = []
    for code, uid in mapping.items():
        hint = hints.get(code)
        name = by_uid.get(uid, '')
        if hint and name and hint.lower() not in name.lower():
            avertissements.append(
                f"{code} — {labels.get(code, code)} : le formulaire sélectionné "
                f"(« {name} ») ne contient pas « {hint} » — vérifiez qu'il "
                f"s'agit bien du bon formulaire."
            )
    return erreurs, avertissements


def _norm_name(s):
    """Normalise un nom d'établissement (referentiel org_unit) pour le
    rapprochement avec le fichier de tirage : retire le suffixe type entre
    parenthèses (ex. « (CSU) », « (EPH (EPHR/EPHD)) ») ajouté par le
    référentiel, espaces multiples, casse. Ne s'applique qu'au nom du
    référentiel — le fichier de tirage n'a pas ce suffixe type (colonne
    séparée), donc pas besoin de la même opération côté tirage (voir
    _norm_tirage_name : un nom comme « EPHR PUBLIC de SAN-PEDRO (Nouveau) »
    a une parenthèse qui fait partie du nom réel, pas un suffixe type — la
    stripper aurait cassé le rapprochement)."""
    if not s:
        return ""
    s = re.sub(r'\s*\([^()]*(\([^()]*\))?[^()]*\)\s*$', '', s).strip()
    return re.sub(r'\s+', ' ', s).upper()


def _norm_tirage_name(s):
    """Normalise un nom d'établissement issu du fichier de tirage — pas de
    parenthèse à retirer ici (voir note dans _norm_name)."""
    if not s:
        return ""
    return re.sub(r'\s+', ' ', s.strip()).upper()


def _etab_type_and_f6_target(label):
    """Devine le type d'établissement à partir du libellé et la cible F6 associée."""
    upper = label.upper()
    if 'CSR-DM' in upper:
        return 'CSR-DM', 2
    if 'CSR-D' in upper or re.search(r'\bCSR\b', upper):
        return 'CSR-D', 1
    if 'EPH' in upper:
        return 'EPH', 3
    if 'CSU' in upper:
        return 'CSU', 3
    return '?', 1


def _load_org_unit(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['choices_etab']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    regions = {}      # code -> {code, nom}
    districts = {}     # code -> {code, nom, region_code}
    etablissements = {}  # code -> {...}
    enqueteurs = {}     # code -> {...}
    superviseurs = {}   # code -> {...}

    for r in rows:
        if not r or not r[0]:
            continue
        list_name, name, label = r[0], r[1], r[2]
        code_id   = r[3] if len(r) > 3 else None
        region_c  = r[4] if len(r) > 4 else None
        district_c = r[5] if len(r) > 5 else None
        enq_c     = r[6] if len(r) > 6 else None

        if list_name == 'Admin_Region_Sanitaire_':
            regions[name] = {'code': name, 'nom': label}

        elif list_name == 'Admin_District_Sanitaire_':
            districts[name] = {'code': name, 'nom': label, 'district_id': code_id, 'region_code': region_c}

        elif list_name == 'Admin_Etablissement_Sante_':
            etype, f6_target = _etab_type_and_f6_target(label)
            etablissements[name] = {
                'code': name,
                'nom_complet': label,
                'nom_normalise': _norm_name(label),
                'etab_id': code_id,
                'region_code': region_c,
                'district_code': district_c,
                'enqueteur_code': enq_c,
                'type': etype,
                'f6_target': f6_target,
                'sig_deces_maternels': 0,  # complété par _apply_tirage_targets
            }

        elif list_name == 'Admin_Enqueteur_':
            enqueteurs[name] = {'code': name, 'nom_complet': label, 'enq_id': code_id,
                                 'region_code': region_c, 'district_code': district_c}

        elif list_name == 'Admin_Superviseur_':
            superviseurs[name] = {'code': name, 'nom_complet': label, 'sup_id': code_id,
                                   'region_code': region_c, 'district_code': district_c}

    return {
        'regions': regions, 'districts': districts, 'etablissements': etablissements,
        'enqueteurs': enqueteurs, 'superviseurs': superviseurs,
    }


def _apply_tirage_targets(ref, tirage_path):
    """Complète le référentiel avec la cible F02 (décès maternels SIG) par
    établissement, en rapprochant les noms entre les deux fichiers sources."""
    wb = openpyxl.load_workbook(tirage_path, read_only=True, data_only=True)
    by_norm_name = {e['nom_normalise']: e for e in ref['etablissements'].values()}

    unmatched = []
    for sheet in wb.sheetnames:
        if not sheet.startswith('Tirage_'):
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for r in rows:
            if not r or not r[2]:
                continue
            etsa_name = r[2].strip()
            sig_deces = r[4] if len(r) > 4 and r[4] else 0
            key = _norm_tirage_name(etsa_name)
            if key in by_norm_name:
                by_norm_name[key]['sig_deces_maternels'] = int(sig_deces or 0)
            else:
                unmatched.append((etsa_name, sig_deces))

    ref['tirage_unmatched'] = unmatched
    return ref


def _apply_local_names(ref, paths):
    """Superpose les noms réels des enquêteurs/superviseurs pour l'affichage,
    à partir du premier fichier trouvé parmi `paths` — jamais versionné sur
    Git (voir .gitignore). Deux emplacements possibles :
      - en local (poste de l'utilisateur) : data/reference/noms_personnel.local.json
      - en ligne (Render) : /etc/secrets/noms_personnel.local.json, injecté
        via un « Secret File » Render — jamais présent dans le dépôt GitHub,
        monté uniquement sur le serveur déployé (voir DEPLOIEMENT_RENDER.md).
    Si aucun des deux n'existe (dépôt public cloné ailleurs, build packagée
    sans le fichier local, Render sans secret configuré) → no-op silencieux,
    le code reste affiché comme avant (comportement sans régression de
    confidentialité)."""
    noms = None
    for path in paths:
        if not path or not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as f:
                noms = json.load(f)
            break
        except Exception:
            continue
    if noms is None:
        return
    for code, nom in noms.get('enqueteurs', {}).items():
        if code in ref['enqueteurs']:
            ref['enqueteurs'][code]['nom_complet'] = nom
    for code, nom in noms.get('superviseurs', {}).items():
        if code in ref['superviseurs']:
            ref['superviseurs'][code]['nom_complet'] = nom


def load(org_unit_path=None, tirage_path=None, use_cache=True):
    """Charge (et met en cache) le référentiel complet à partir des fichiers Excel fournis
    (par défaut : data/reference/org_unit.xlsx et data/reference/tirage_etablissements.xlsx)."""
    org_unit_path = org_unit_path or DEFAULT_ORG_UNIT_PATH
    tirage_path = tirage_path if tirage_path is not None else DEFAULT_TIRAGE_PATH
    cache_key = (org_unit_path, tirage_path)
    if use_cache and cache_key in _cache:
        return _cache[cache_key]

    ref = _load_org_unit(org_unit_path)
    if tirage_path:
        _apply_tirage_targets(ref, tirage_path)
    _apply_local_names(ref, [DEFAULT_LOCAL_NAMES_PATH, RENDER_SECRET_NAMES_PATH])

    _cache[cache_key] = ref
    return ref


def clear_cache():
    _cache.clear()


def target_for(ref, form_code, etablissement_code=None, district_code=None):
    """Cible attendue pour un formulaire donné, au niveau établissement ou
    district — dérivée du registre des formulaires (modules/forms_registry.py)
    plutôt que de règles codées en dur par formulaire. Renvoie None pour un
    formulaire inconnu/désactivé du registre."""
    form = forms_registry.get(form_code)
    if form is None:
        return None
    if form['grain'] == 'district':
        return forms_registry.target_for_district(form, ref, district_code)
    return forms_registry.target_for_etablissement(form, ref, etablissement_code)


def is_floor(form_code):
    """Vrai si la cible de ce formulaire est un plancher (le dépasser est
    normal, jamais une anomalie « à vérifier ») — motif F07 par défaut."""
    form = forms_registry.get(form_code)
    return bool(form and forms_registry.is_floor_rule(form))


def national_targets(ref):
    """Cibles nationales agrégées par formulaire actif — pour la vue nationale."""
    return {
        f['code']: forms_registry.national_target(f, ref)
        for f in forms_registry.all_forms(include_inactive=False)
    }
