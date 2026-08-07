import os, json, uuid, re, datetime, io, csv
import pandas as pd
from flask import (Flask, render_template, request, session,
                   redirect, url_for, jsonify, send_file, flash, Response)
from werkzeug.utils import secure_filename
from config import Config
from modules.data_loader import (load_excel, load_excel_sheets, clean_kobo_dataframe,
                                  get_var_types, summarize_dataframe, get_filtered_vars)
from modules.descriptive import tris_a_plat, statistiques_continues, analyse_binaire_groupe
from modules.crosstabs import tableau_croise, tableau_croise_dynamique
from modules.multi_survey import (merge_surveys, variable_coverage,
                                    common_variables, phakts_radical,
                                    is_phakts_coded,
                                    compare_categorical_by_survey,
                                    compare_continuous_by_survey,
                                    auto_analyze)
from modules.multivariate import run_pca, run_mca, run_clustering, run_ca, select_viable_variables
from modules.report_generator import generate_pdf_report, generate_word_report
from modules.comments import (auto_comment_categorical, auto_comment_continuous,
                              auto_comment_crosstab, auto_comment_binary_group)
from modules.raw_analysis import (systematic_analysis, descriptive_summary,
                                   data_quality_gauge, composition_chart,
                                   distributions_chart, overview_stats,
                                   missing_bar_chart, missing_heatmap, correlation_matrix)

# Import sécurisé du connecteur KoboToolbox
# (nécessite le package 'requests' — installé via : venv/bin/pip install requests)
try:
    from modules.kobo_connector import (validate_token, list_assets,
                                         get_asset_info, load_data as kobo_load_data,
                                         deploy_xlsform)
    KOBO_AVAILABLE = True
except ImportError:
    KOBO_AVAILABLE = False
    def validate_token(*a, **k): return {"valid": False, "error": "Package 'requests' manquant. Lancez : venv/bin/pip install requests"}
    def list_assets(*a, **k):    return {"success": False, "error": "Package 'requests' manquant."}
    def get_asset_info(*a, **k): return {"success": False, "error": "Package 'requests' manquant."}
    def kobo_load_data(*a, **k): return {"success": False, "error": "Package 'requests' manquant."}
    def deploy_xlsform(*a, **k): return {"success": False, "error": "Package 'requests' manquant."}

from modules import kobo_sync
from modules import kobo_track
from modules import reference_data as ref_data
from modules import completeness as cp
from modules import projets as proj
from modules import tendance
from modules import completude_report
from modules import form_mapping
from modules import accounts
from modules import activity_log
from modules import forms_registry
from modules import ai_form_assist
from modules.collecte_monitor import load_state, save_state, append_sync_event, build_dashboard_metrics, build_collecte_views

app = Flask(__name__)
app.config.from_object(Config)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['REPORTS_FOLDER'], exist_ok=True)

@app.context_processor
def inject_current_year():
    """Année courante (calendrier grégorien) injectée dans tous les templates —
    évite de coder en dur « SPAD 2026 » dans le bandeau (voir base.html)."""
    return {'current_year': datetime.datetime.now().year}


if (os.environ.get('ANALYZER_PASSWORD_ADMIN') or os.environ.get('ANALYZER_PASSWORD_INVITE')) \
        and not os.environ.get('SECRET_KEY'):
    print(
        "\n⚠️  ANALYZER_PASSWORD_ADMIN/ANALYZER_PASSWORD_INVITE défini(s) mais SECRET_KEY "
        "ne l'est pas — les sessions utilisent la clé par défaut du code source (non "
        "secrète). Définissez SECRET_KEY (valeur aléatoire) dans les variables "
        "d'environnement avant toute mise en production.\n"
    )


# ── Gestionnaire d'erreur 500 — affiche le traceback complet ─────────────────
import traceback

@app.errorhandler(500)
def internal_error(e):
    tb = traceback.format_exc()
    html = f"""
    <html><head><title>SPAD — Erreur 500</title>
    <style>
      body {{font-family:monospace;background:#1C2833;color:#FDFEFE;padding:30px}}
      h2   {{color:#E74C3C}}
      pre  {{background:#17202A;padding:20px;border-radius:8px;
             border-left:4px solid #E74C3C;overflow-x:auto;font-size:13px;line-height:1.6}}
      .tip {{background:#1A5276;padding:14px 18px;border-radius:8px;
             margin-top:20px;color:#AED6F1;font-size:13px}}
    </style></head><body>
    <h2>⚠ Erreur interne 500</h2>
    <pre>{tb}</pre>
    <div class="tip">
      📋 Copiez ce traceback complet et envoyez-le pour correction immédiate.
    </div>
    </body></html>
    """
    return html, 500


@app.route('/favicon.ico')
def favicon():
    """Redirige vers le vrai logo SPAD (static/img/spad_favicon.png) — certains
    navigateurs demandent /favicon.ico même quand un <link rel="icon"> existe."""
    return redirect(url_for('static', filename='img/spad_favicon.png'))


# ─── Authentification (3 rôles) ────────────────────────────────────────────────
#
# En local (poste de l'utilisateur), l'accès est déjà restreint par le système
# d'exploitation : seule la personne devant la machine peut atteindre 127.0.0.1.
# Dès que l'app est hébergée sur une URL publique (ex. Render), ce n'est plus
# vrai — d'où ce verrou, activé uniquement si un mot de passe Admin ou Invité
# est défini (donc sans impact sur l'usage desktop local existant tant
# qu'aucune des deux variables n'est positionnée).
#
#   - ANALYZER_PASSWORD_ADMIN  -> rôle 'admin'  : accès complet + gestion des
#     comptes Data (autoriser/bloquer) et consultation du journal d'activité.
#     Mot de passe partagé (un seul admin, ou peu), comme avant.
#   - Compte individuel (identifiant + mot de passe, voir modules/accounts.py)
#     -> rôle 'data' : accès complet, comme avant — mais chaque utilisateur a
#     désormais son propre compte plutôt qu'un mot de passe partagé, pour que
#     ses actions soient attribuables (voir modules/activity_log.py). Créé
#     par auto-inscription (/register), inactif tant qu'un admin ne l'a pas
#     approuvé. ANCIEN ANALYZER_PASSWORD (mot de passe Data partagé) retiré —
#     les comptes existants doivent migrer vers un compte individuel.
#   - ANALYZER_PASSWORD_INVITE -> rôle 'invite' : lecture seule, restreint au
#     tableau de bord Complétude nationale / par district & établissement /
#     performances superviseurs / performances enquêteurs — voir
#     INVITE_ALLOWED_ENDPOINTS ci-dessous. Pas d'export, pas de recalcul, pas
#     d'accès à la connexion KoboToolbox ni aux autres modules d'analyse.
ANALYZER_PASSWORD_ADMIN = os.environ.get('ANALYZER_PASSWORD_ADMIN', '').strip()
ANALYZER_PASSWORD_INVITE = os.environ.get('ANALYZER_PASSWORD_INVITE', '').strip()

# Endpoints accessibles au rôle 'invite'. Volontairement une liste blanche
# explicite (plutôt qu'une liste noire) : toute nouvelle route ajoutée plus
# tard reste bloquée pour ce rôle par défaut, jusqu'à décision explicite de
# l'ouvrir.
INVITE_ALLOWED_ENDPOINTS = {
    'login', 'logout', 'static', 'favicon',
    'completude',
    'completude_districts', 'completude_district_detail', 'completude_etablissement_detail',
    'completude_superviseurs', 'completude_superviseur_detail',
    'completude_enqueteurs', 'completude_enqueteur_detail',
}

# Endpoints réservés au rôle 'admin' — bloqués pour 'data' et 'invite'.
ADMIN_ONLY_ENDPOINTS = {
    'admin_users', 'admin_user_approve', 'admin_user_block', 'admin_activity',
    'admin_forms', 'admin_form_activate', 'admin_form_deactivate', 'admin_form_add',
}


@app.before_request
def require_login():
    if not ANALYZER_PASSWORD_ADMIN and not ANALYZER_PASSWORD_INVITE:
        return  # aucun mot de passe configuré : gate désactivée (usage desktop local)
    if request.endpoint in ('login', 'register', 'static', 'favicon'):
        return
    if request.path.startswith('/static/'):
        return
    if not session.get('authenticated'):
        return redirect(url_for('login', next=request.path))
    role = session.get('role')
    if role != 'admin' and request.endpoint in ADMIN_ONLY_ENDPOINTS:
        flash("Cette page est réservée à l'administrateur.", 'warning')
        return redirect(url_for('completude') if role == 'invite' else url_for('index'))
    if role == 'invite' and request.endpoint not in INVITE_ALLOWED_ENDPOINTS:
        flash("Cette page n'est pas accessible avec un accès Invité.", 'warning')
        return redirect(url_for('completude'))


@app.after_request
def log_data_activity(response):
    """Journalise chaque action d'un utilisateur 'data' — voir
    modules/activity_log.py. Jamais 'invite' (lecture seule, rien à
    auditer) ni 'admin'. Best-effort : ne bloque jamais la réponse réelle."""
    if session.get('role') == 'data' and request.endpoint not in (None, 'static', 'favicon'):
        activity_log.record(session.get('username', '?'), 'data', request.method, request.path)
    return response


@app.before_request
def auto_kobo_connect():
    """Si un jeton KoboToolbox serveur est configuré (KOBO_API_TOKEN) et que
    l'utilisateur Data courant n'a pas encore de connexion Kobo en session,
    connecte automatiquement — pour que personne n'ait à saisir de token pour
    parcourir/charger un formulaire Kobo (cliquer « KoboToolbox » affiche
    directement la liste des formulaires). Un utilisateur peut toujours se
    connecter avec son propre token via /kobo/connect pour utiliser un autre
    compte Kobo que celui configuré sur le serveur.
    Portée : rôles 'data' et 'admin' — le rôle 'invite' n'accède à aucune
    page qui en a besoin (voir INVITE_ALLOWED_ENDPOINTS)."""
    if session.get('role') not in ('data', 'admin') or session.get('kobo_token'):
        return
    server_token = (os.environ.get('KOBO_API_TOKEN') or '').strip()
    if not server_token:
        return
    server_instance = (os.environ.get('KOBO_INSTANCE') or '').strip() or None
    result = validate_token(server_token, custom_instance=server_instance)
    if result.get('valid'):
        session['kobo_token']    = server_token
        session['kobo_username'] = result.get('username', '')
        session['kobo_instance'] = result.get('instance', '')
        kobo_track.resume(server_token, result.get('instance'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if not ANALYZER_PASSWORD_ADMIN and not ANALYZER_PASSWORD_INVITE:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        entered = request.form.get('password', '')
        access  = request.form.get('access', '')
        # Chaque onglet de connexion (Admin / Data / Invité) ne valide que
        # son propre mécanisme — pas de mélange entre les mots de passe
        # partagés (Admin, Invité) et les comptes individuels (Data).
        if access == 'admin' and ANALYZER_PASSWORD_ADMIN and entered == ANALYZER_PASSWORD_ADMIN:
            session['authenticated'] = True
            session['role'] = 'admin'
            session.permanent = True
            return redirect(request.args.get('next') or url_for('index'))
        elif access == 'data':
            status, real_username = accounts.verify_login(request.form.get('username', ''), entered)
            if status == accounts.STATUS_APPROVED:
                session['authenticated'] = True
                session['role'] = 'data'
                session['username'] = real_username
                session.permanent = True
                return redirect(request.args.get('next') or url_for('index'))
            elif status == accounts.STATUS_PENDING:
                error = "Ce compte est en attente de validation par un administrateur."
            elif status == accounts.STATUS_BLOCKED:
                error = "Ce compte a été bloqué. Contactez un administrateur."
            else:
                error = 'Identifiant ou mot de passe incorrect.'
        elif access == 'invite' and ANALYZER_PASSWORD_INVITE and entered == ANALYZER_PASSWORD_INVITE:
            session['authenticated'] = True
            session['role'] = 'invite'
            session.permanent = True
            return redirect(url_for('completude'))
        if not error:
            error = 'Mot de passe incorrect.'
    return render_template('login.html', error=error)


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Auto-inscription d'un compte Data — reste en attente jusqu'à
    validation par un administrateur (voir /admin/users)."""
    if not ANALYZER_PASSWORD_ADMIN and not ANALYZER_PASSWORD_INVITE:
        return redirect(url_for('index'))
    error = None
    success = False
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if password != confirm:
            error = "Les deux mots de passe ne correspondent pas."
        else:
            ok, err = accounts.create_pending(username, password)
            if ok:
                success = True
            else:
                error = err
    return render_template('register.html', error=error, success=success)


@app.route('/logout')
def logout():
    session.pop('authenticated', None)
    session.pop('role', None)
    session.pop('username', None)
    return redirect(url_for('login'))


# ─── Administration (rôle 'admin' uniquement) ──────────────────────────────────

@app.route('/admin/users')
def admin_users():
    return render_template('admin_users.html', users=accounts.list_users())


@app.route('/admin/users/<username>/approve', methods=['POST'])
def admin_user_approve(username):
    if accounts.set_status(username, accounts.STATUS_APPROVED):
        flash(f"Compte « {username} » autorisé.", 'success')
    else:
        flash('Compte introuvable.', 'danger')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<username>/block', methods=['POST'])
def admin_user_block(username):
    if accounts.set_status(username, accounts.STATUS_BLOCKED):
        flash(f"Compte « {username} » bloqué.", 'warning')
    else:
        flash('Compte introuvable.', 'danger')
    return redirect(url_for('admin_users'))


@app.route('/admin/activity')
def admin_activity():
    return render_template('admin_activity.html', events=activity_log.list_events())


@app.route('/admin/forms')
def admin_forms():
    """Registre des formulaires suivis (modules/forms_registry.py) — activer/
    désactiver les 7 formulaires SPAD historiques, en ajouter d'autres."""
    return render_template(
        'admin_forms.html',
        forms=forms_registry.all_forms(),
        rule_types=forms_registry.RULE_TYPES,
        rule_type_labels=forms_registry.RULE_TYPE_LABELS,
    )


@app.route('/admin/forms/<code>/activate', methods=['POST'])
def admin_form_activate(code):
    if forms_registry.set_active(code, True):
        flash(f"Formulaire « {code} » activé.", 'success')
    else:
        flash('Formulaire introuvable.', 'danger')
    return redirect(url_for('admin_forms'))


@app.route('/admin/forms/<code>/deactivate', methods=['POST'])
def admin_form_deactivate(code):
    if forms_registry.set_active(code, False):
        flash(f"Formulaire « {code} » désactivé — il n'apparaîtra plus dans Suivi ni Complétude.", 'warning')
    else:
        flash('Formulaire introuvable.', 'danger')
    return redirect(url_for('admin_forms'))


@app.route('/admin/forms/add', methods=['POST'])
def admin_form_add():
    rtype = request.form.get('rule_type', '')
    params = {}
    if rtype in ('fixed_per_etablissement', 'fixed_per_district'):
        try:
            params = {'n': int(request.form.get('rule_n', ''))}
        except ValueError:
            flash('La valeur numérique de la règle de cible est invalide.', 'danger')
            return redirect(url_for('admin_forms'))
    elif rtype in ('etab_field_positive', 'floor_sum_district_field'):
        field = (request.form.get('rule_field') or '').strip()
        if not field:
            flash('Le nom du champ du référentiel est obligatoire pour ce type de règle.', 'danger')
            return redirect(url_for('admin_forms'))
        params = {'field': field}

    ok, err = forms_registry.add_form(
        code=request.form.get('code', ''),
        label=request.form.get('label', ''),
        name_hint=request.form.get('name_hint', ''),
        grain=request.form.get('grain', ''),
        actor=request.form.get('actor', ''),
        etab_field=request.form.get('etab_field', ''),
        district_field=request.form.get('district_field', ''),
        target_rule={'type': rtype, 'params': params},
    )
    if ok:
        flash(f"Formulaire ajouté au registre.", 'success')
    else:
        flash(err, 'danger')
    return redirect(url_for('admin_forms'))


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def get_dataframe(sheet_key='data_path'):
    path = session.get(sheet_key)
    if path and os.path.exists(path):
        return pd.read_excel(path)
    return None


def _collecte_state_path():
    state_path = session.get('collecte_state_path')
    if state_path:
        return state_path
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'collecte_state.json')
    session['collecte_state_path'] = path
    return path


def _save_dataframe_to_session(df, name):
    fname = f"kobo_{uuid.uuid4().hex[:8]}.xlsx"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    df.to_excel(save_path, index=False, engine='openpyxl')
    session['data_path'] = save_path
    session['data_meta'] = summarize_dataframe(df)
    session['data_meta']['source'] = 'KoboToolbox'
    session['data_meta']['name'] = name
    session['original_filename'] = f"{name}.xlsx"
    session.pop('child_path', None)
    return save_path


@app.route('/collecte/dashboard')
def collecte_dashboard():
    state_path = _collecte_state_path()
    state = load_state(state_path)
    current_count = None
    meta = session.get('data_meta') or {}
    if meta.get('n_obs') is not None:
        current_count = int(meta['n_obs'])
    metrics = build_dashboard_metrics(state, current_count=current_count)
    sync_status = kobo_sync.status() if hasattr(kobo_sync, 'status') else {}
    views = build_collecte_views(state, current_count=current_count, data_meta=meta, sync_status=sync_status)
    return render_template(
        'collecte_dashboard.html',
        received=metrics['received'],
        cible=metrics['cible'],
        taux=metrics['taux'],
        active_alerts=metrics['active_alerts'],
        evolution=metrics['evolution'],
        zones=metrics['zones'],
        kobo_connected=bool(session.get('kobo_token')),
        kobo_user=session.get('kobo_username', 'KoboToolbox'),
        last_sync=state.get('last_sync_at') or 'Aucune synchronisation',
        history=metrics['history'],
        daily_rate=metrics.get('daily_rate', 0),
        daily_target=metrics.get('daily_target', 0),
        sync_status=views['sync_status'],
        stratified_summary=views['stratified_summary'],
        interpretation=views['interpretation'],
    )


@app.route('/collecte/sync')
def collecte_sync():
    state_path = _collecte_state_path()
    state = load_state(state_path)
    history = []
    for item in state.get('history', []):
        history.append({
            'date': item.get('timestamp') or state.get('last_sync_at') or 'Non disponible',
            'form': item.get('form_name') or state.get('last_form_name') or 'Kobo',
            'count': item.get('count', 0),
            'status': item.get('status', 'inconnu').capitalize(),
            'badge_class': 'bg-success' if item.get('status') == 'réussi' else 'bg-warning',
        })
    assets = []
    selected_uid = session.get('kobo_uid')
    token = session.get('kobo_token')
    if token:
        assets_result = list_assets(token, instance=session.get('kobo_instance'))
        if assets_result.get('success'):
            assets = assets_result.get('assets', [])
            if not selected_uid and assets:
                selected_uid = assets[0].get('uid')
    return render_template(
        'collecte_sync.html',
        kobo_connected=bool(session.get('kobo_token')),
        kobo_user=session.get('kobo_username', 'KoboToolbox'),
        kobo_instance=session.get('kobo_instance') or 'KoboToolbox',
        sync_history=history,
        last_sync_status=state.get('last_sync_status', 'inconnu'),
        last_sync_count=state.get('last_sync_count', 0),
        assets=assets,
        selected_uid=selected_uid,
        target=state.get('target', 0),
        auto_interval=300,
        sync_status=kobo_sync.status() if hasattr(kobo_sync, 'status') else {},
    )


@app.route('/collecte/sync/run', methods=['POST'])
def collecte_sync_run():
    token = session.get('kobo_token')
    uid = (request.form.get('uid') or session.get('kobo_uid') or '').strip()
    instance = session.get('kobo_instance')
    name = session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    if not token or not uid:
        flash('Connectez-vous d’abord à KoboToolbox pour synchroniser les données.', 'warning')
        return redirect(url_for('collecte_sync'))
    target_raw = (request.form.get('target') or '').strip()
    target = int(target_raw) if target_raw else None
    state_path = _collecte_state_path()
    state = load_state(state_path)
    result = kobo_load_data(token, uid, instance=instance)
    if not result.get('success'):
        state = append_sync_event(state, state_path, form_name=name, count=state.get('last_sync_count', 0), status='erreur', target=target or state.get('target'))
        flash(f"Synchronisation échouée : {result.get('error', 'Erreur inconnue')}", 'danger')
        return redirect(url_for('collecte_sync'))
    df = result['df']
    asset_info = get_asset_info(token, uid, instance=instance)
    name = asset_info.get('name') or session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    _save_dataframe_to_session(df, name)
    session['kobo_uid'] = uid
    session['kobo_asset_name'] = name
    state = append_sync_event(state, state_path, form_name=name, count=int(result.get('n_obs', 0)), status='réussi', target=target or state.get('target'))
    session['collecte_state_path'] = state_path
    flash(f"Synchronisation réussie — {result.get('n_obs')} soumissions chargées depuis {name}.", 'success')
    return redirect(url_for('collecte_sync'))


@app.route('/collecte/sync/auto', methods=['POST'])
def collecte_sync_auto():
    token = session.get('kobo_token')
    uid = session.get('kobo_uid')
    instance = session.get('kobo_instance')
    name = session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    if not token or not uid:
        flash('Connectez-vous d’abord à KoboToolbox pour activer la synchronisation automatique.', 'warning')
        return redirect(url_for('collecte_sync'))
    interval = int(request.form.get('interval_seconds', '300') or 300)
    baseline = int((session.get('data_meta') or {}).get('n_obs', 0) or 0)
    kobo_sync.start(token, uid, instance, name, interval, baseline)
    flash('Synchronisation automatique démarrée.', 'success')
    return redirect(url_for('collecte_sync'))


@app.route('/collecte/refresh', methods=['POST'])
def collecte_refresh():
    token = session.get('kobo_token')
    uid = session.get('kobo_uid')
    instance = session.get('kobo_instance')
    if not token or not uid:
        flash('Aucun formulaire Kobo actif. Connectez-vous d’abord.', 'warning')
        return redirect(url_for('collecte_sync'))

    state_path = _collecte_state_path()
    state = load_state(state_path)
    result = kobo_load_data(token, uid, instance=instance)
    if not result.get('success'):
        state = append_sync_event(state, state_path, form_name=session.get('kobo_asset_name', 'Formulaire KoboToolbox'), count=state.get('last_sync_count', 0), status='erreur', target=state.get('target'))
        flash(f"Rafraîchissement échoué : {result.get('error', 'Erreur inconnue')}", 'danger')
        return redirect(url_for('collecte_dashboard'))

    df = result['df']
    asset_info = get_asset_info(token, uid, instance=instance)
    name = asset_info.get('name') or session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    _save_dataframe_to_session(df, name)
    session['kobo_uid'] = uid
    session['kobo_asset_name'] = name
    state = append_sync_event(state, state_path, form_name=name, count=int(result.get('n_obs', 0)), status='réussi', target=state.get('target'))
    flash(f"Données rafraîchies — {result.get('n_obs')} soumissions disponibles.", 'success')
    return redirect(url_for('collecte_dashboard'))


@app.route('/collecte/geographique')
def collecte_geographique():
    state_path = _collecte_state_path()
    state = load_state(state_path)
    current_count = None
    meta = session.get('data_meta') or {}
    if meta.get('n_obs') is not None:
        current_count = int(meta['n_obs'])
    sync_status = kobo_sync.status() if hasattr(kobo_sync, 'status') else {}
    views = build_collecte_views(state, current_count=current_count, data_meta=meta, sync_status=sync_status)
    return render_template('collecte_geographique.html', geo_items=views['geo_items'], sync_status=views['sync_status'])


@app.route('/collecte/equipes')
def collecte_equipes():
    state_path = _collecte_state_path()
    state = load_state(state_path)
    current_count = None
    meta = session.get('data_meta') or {}
    if meta.get('n_obs') is not None:
        current_count = int(meta['n_obs'])
    sync_status = kobo_sync.status() if hasattr(kobo_sync, 'status') else {}
    views = build_collecte_views(state, current_count=current_count, data_meta=meta, sync_status=sync_status)
    return render_template('collecte_equipes.html', teams=views['teams'], sync_status=views['sync_status'])


@app.route('/collecte/alertes')
def collecte_alertes():
    state_path = _collecte_state_path()
    state = load_state(state_path)
    current_count = None
    meta = session.get('data_meta') or {}
    if meta.get('n_obs') is not None:
        current_count = int(meta['n_obs'])
    sync_status = kobo_sync.status() if hasattr(kobo_sync, 'status') else {}
    views = build_collecte_views(state, current_count=current_count, data_meta=meta, sync_status=sync_status)
    return render_template('collecte_alertes.html', alerts=views['alerts'], sync_status=views['sync_status'])


@app.route('/collecte/qualite')
def collecte_qualite():
    meta = session.get('data_meta') or {}
    state_path = _collecte_state_path()
    state = load_state(state_path)
    sync_status = kobo_sync.status() if hasattr(kobo_sync, 'status') else {}
    views = build_collecte_views(state, current_count=meta.get('n_obs'), data_meta=meta, sync_status=sync_status)
    return render_template('collecte_qualite.html', quality_items=views['quality_items'], sync_status=views['sync_status'])


def detect_binary_groups(df: pd.DataFrame) -> list[dict]:
    """
    Détecte les groupes de variables binaires KoBoToolbox (questions à choix multiples).
    Ex: "Raison/Peur" "Raison/Cout" → groupe "Raison"
    """
    var_types = get_var_types(df)
    bin_cols = [c for c, t in var_types.items() if t == 'binaire']
    groups = {}
    for col in bin_cols:
        if '/' in col:
            parent = col.rsplit('/', 1)[0].strip()
            groups.setdefault(parent, []).append(col)
    # Keep only groups with ≥ 2 items
    return [{'parent': p, 'vars': v} for p, v in groups.items() if len(v) >= 2]


# ─── User comments helpers ──────────────────────────────────────────────────
def _get_user_comment(section: str, var: str) -> str:
    """Lit un commentaire libre stocké en session."""
    return (session.get('user_comments', {}) or {}).get(f"{section}:{var}", '')


@app.route('/api/save-comment', methods=['POST'])
def save_comment():
    """Stocke un commentaire libre dans la session (par section + variable)."""
    data = request.get_json(silent=True) or request.form
    section = (data.get('section') or '').strip()
    var = (data.get('var') or '').strip()
    text = (data.get('text') or '').strip()
    if not section or not var:
        return jsonify({'error': 'section et var requis'}), 400
    comments = dict(session.get('user_comments', {}) or {})
    key = f"{section}:{var}"
    if text:
        comments[key] = text[:2000]   # safety cap
    else:
        comments.pop(key, None)
    session['user_comments'] = comments
    return jsonify({'ok': True})


# ─── HOME ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    has_data = bool(session.get('data_path') and os.path.exists(session.get('data_path', '')))
    meta = session.get('data_meta', {})
    sheets = session.get('sheet_names', [])
    return render_template('index.html', has_data=has_data, meta=meta, sheets=sheets)


# ─── ANALYSE DE DONNÉES (point d'entrée unique : Kobo ou fichier local) ───────

@app.route('/analyse-donnees')
def analyse_donnees():
    has_data = bool(session.get('data_path') and os.path.exists(session.get('data_path', '')))
    return render_template(
        'analyse_donnees.html',
        kobo_connected=bool(session.get('kobo_token')),
        has_data=has_data, meta=session.get('data_meta', {}),
    )


# ─── UPLOAD ───────────────────────────────────────────────────────────────────

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(request.url)
        f = request.files['file']
        if f.filename == '':
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(request.url)
        if f and allowed_file(f.filename):
            fname = str(uuid.uuid4()) + '_' + secure_filename(f.filename)
            raw_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            f.save(raw_path)
            try:
                # Load all sheets
                all_sheets = load_excel_sheets(raw_path)
                sheet_names = list(all_sheets.keys())
                session['sheet_names'] = sheet_names
                session['original_filename'] = f.filename

                # Main sheet (first)
                main_name = sheet_names[0]
                df_raw = all_sheets[main_name]
                df_clean = clean_kobo_dataframe(df_raw)

                # Save cleaned main sheet
                base = fname.rsplit('.', 1)[0]
                clean_path = os.path.join(
                    app.config['UPLOAD_FOLDER'], f'clean_{base}.xlsx')
                df_clean.to_excel(clean_path, index=False, engine='openpyxl')
                session['data_path'] = clean_path
                session['data_meta'] = summarize_dataframe(df_clean)
                session['data_meta']['sheet_name'] = main_name

                # Save child sheet if exists
                if len(sheet_names) > 1:
                    df_child = all_sheets[sheet_names[1]]
                    df_child_clean = clean_kobo_dataframe(df_child)
                    child_path = os.path.join(
                        app.config['UPLOAD_FOLDER'], f'child_{base}.xlsx')
                    df_child_clean.to_excel(child_path, index=False, engine='openpyxl')
                    session['child_path'] = child_path
                    session['data_meta']['has_child_sheet'] = True
                    session['data_meta']['child_sheet_name'] = sheet_names[1]
                    session['data_meta']['n_child'] = len(df_child_clean)
                else:
                    session['data_meta']['has_child_sheet'] = False

                flash(
                    f'Fichier chargé avec succès — {len(df_clean)} ménages × '
                    f'{len(df_clean.columns)} variables analysables.',
                    'success')
                return redirect(url_for('data_preview'))
            except Exception as e:
                flash(f'Erreur lors du chargement : {str(e)}', 'danger')
        else:
            flash('Format non supporté. Utilisez .xlsx ou .xls', 'danger')
    return render_template('upload.html')


def _phakts_libelle_from_code(col: str) -> str:
    """
    Reconstruit un libellé lisible à partir d'un nom de colonne PHAKTS.
    Ex.: 'Statut_Matrimonial__X!1*1Statut_Matrimonial_' → 'Statut Matrimonial'.
    Si le nom n'est pas codifié PHAKTS, renvoie le nom sans suffixe technique.
    """
    if not col:
        return ''
    s = str(col).strip()
    # Retire le suffixe PHAKTS (__TYPE...)
    base = s.split('__', 1)[0] if '__' in s else s
    # Retire un éventuel suffixe binaire « parent/enfant »
    base = base.split('/')[-1]
    return base.replace('_', ' ').strip()


def _extract_modalites(series, typ: str, max_items: int = 20) -> str:
    """
    Extrait les modalités/valeurs typiques d'une variable, sous une forme
    compacte adaptée à un questionnaire imprimé.
      - catégorielle / binaire : « Modalité1, Modalité2, … » (max 20)
      - continue : « plage : min – max »
      - date : « plage : min – max »
      - texte_libre : « (texte libre) »
    """
    try:
        s = series.dropna()
        if s.empty:
            return ''
        if typ in ('categorielle', 'binaire'):
            # binaire 0/1 → Oui/Non
            try:
                import pandas as pd
                num = pd.to_numeric(s, errors='raise')
                if set(num.unique()).issubset({0, 1, 0.0, 1.0}):
                    return 'Oui ; Non'
            except Exception:
                pass
            vals = [str(v).strip() for v in s.astype(str).value_counts().index[:max_items]]
            extra = '' if len(set(s.astype(str))) <= max_items else f' (+{len(set(s.astype(str)))-max_items} autres)'
            return ' ; '.join(vals) + extra
        if typ == 'continue':
            import pandas as pd
            num = pd.to_numeric(s, errors='coerce').dropna()
            if num.empty:
                return ''
            return f'plage : {round(float(num.min()), 2)} – {round(float(num.max()), 2)}'
        if typ == 'date':
            try:
                import pandas as pd
                dt = pd.to_datetime(s, errors='coerce').dropna()
                if dt.empty:
                    return ''
                return f'plage : {dt.min().strftime("%Y-%m-%d")} – {dt.max().strftime("%Y-%m-%d")}'
            except Exception:
                return ''
        return '(texte libre)'
    except Exception:
        return ''


def _build_var_types_table(df) -> list:
    """
    Construit la liste { variable, type, libelle, modalites, n_obs, n_manq, %manq }
    pour l'affichage et l'export XLSX. La colonne « libellé » privilégie celui
    saisi par l'utilisateur (session['var_labels']) puis le radical PHAKTS.
    """
    var_types = get_var_types(df)
    user_labels = session.get('var_labels', {}) or {}
    rows = []
    n = len(df)
    for i, (col, typ) in enumerate(var_types.items(), start=1):
        n_miss = int(df[col].isna().sum())
        auto_lib = _phakts_libelle_from_code(col)
        custom_lib = user_labels.get(col, '').strip()
        rows.append({
            'idx': i,
            'variable': col,
            'libelle': custom_lib or auto_lib,
            'libelle_auto': auto_lib,
            'is_custom': bool(custom_lib),
            'modalites': _extract_modalites(df[col], typ),
            'type': typ,
            'n_obs': int(n - n_miss),
            'n_manq': n_miss,
            'pct_manq': round(n_miss / n * 100, 1) if n else 0,
        })
    return rows


@app.route('/api/var-label', methods=['POST'])
def save_var_label():
    """Sauvegarde un libellé personnalisé pour une variable, en session."""
    data = request.get_json(silent=True) or request.form
    var = (data.get('variable') or '').strip()
    label = (data.get('label') or '').strip()
    if not var:
        return jsonify({'error': "variable requis"}), 400
    labels = dict(session.get('var_labels', {}) or {})
    if label:
        labels[var] = label[:500]
    else:
        labels.pop(var, None)
    session['var_labels'] = labels
    return jsonify({'ok': True, 'variable': var, 'label': label})


@app.route('/data/preview')
def data_preview():
    df = get_dataframe()
    if df is None:
        flash('Veuillez d\'abord importer un fichier.', 'warning')
        return redirect(url_for('upload'))
    var_types = get_var_types(df)
    var_table = _build_var_types_table(df)
    bin_groups = detect_binary_groups(df)
    preview = df.head(20).to_html(
        classes='table table-sm table-striped table-hover',
        border=0, index=False, na_rep='—')
    meta = session.get('data_meta', {})

    # Child sheet preview
    child_preview = None
    child_meta = {}
    if session.get('child_path') and os.path.exists(session['child_path']):
        df_child = pd.read_excel(session['child_path'])
        child_preview = df_child.head(20).to_html(
            classes='table table-sm table-striped table-hover',
            border=0, index=False, na_rep='—')
        child_meta = {'n': len(df_child), 'cols': len(df_child.columns)}

    return render_template('data_preview.html',
                           preview=preview, columns=df.columns.tolist(),
                           var_types=var_types, var_table=var_table,
                           meta=meta, shape=df.shape,
                           bin_groups=bin_groups,
                           child_preview=child_preview, child_meta=child_meta)


@app.route('/data/var-types.xlsx')
def export_var_types_xlsx():
    """
    Export XLSX = vrai questionnaire reconstruit à partir des données.
    Feuille 1 « Questionnaire » : N° / Variable / Question / Type / Modalités attendues /
    Obs. valides / Manquantes / % manq.
    Feuille 2 « Métadonnées » : nom du fichier, taille, date d'export.
    """
    df = get_dataframe()
    if df is None:
        flash("Veuillez d'abord importer un fichier.", 'warning')
        return redirect(url_for('upload'))
    rows = _build_var_types_table(df)
    TYPE_LABEL = {
        'categorielle': 'Choix unique / multiple',
        'continue':     'Numérique',
        'binaire':      'Booléen (Oui/Non)',
        'date':         'Date',
        'texte_libre':  'Texte libre',
    }
    table = pd.DataFrame([{
        'N°': r['idx'],
        'Variable (code)': r['variable'],
        'Question / Libellé': r['libelle'],
        'Type': TYPE_LABEL.get(r['type'], r['type']),
        'Modalités attendues': r['modalites'],
        'Obs. valides': r['n_obs'],
        'Manquantes': r['n_manq'],
        '% manquantes': r['pct_manq'],
    } for r in rows])

    meta = pd.DataFrame([
        ['Fichier source',  session.get('original_filename', '')],
        ['Nb. observations', int(len(df))],
        ['Nb. variables',    int(df.shape[1])],
        ['Date d\'export',   pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')],
    ], columns=['Indicateur', 'Valeur'])

    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        table.to_excel(writer, index=False, sheet_name='Questionnaire')
        meta.to_excel(writer, index=False, sheet_name='Métadonnées')

        # Mise en forme feuille « Questionnaire »
        ws = writer.sheets['Questionnaire']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1A5276')
        thin = Side(border_style='thin', color='BDC3C7')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for c in col_cells:
                if c.value is not None:
                    L = len(str(c.value))
                    if L > max_len: max_len = L
                if c.row > 1:
                    c.alignment = Alignment(vertical='top', wrap_text=True)
                    c.border = border
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(12, max_len + 2), 60)
        ws.row_dimensions[1].height = 30

        # Feuille Métadonnées : entête simple
        ws2 = writer.sheets['Métadonnées']
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E8EAF0')

    buf.seek(0)
    fname = (session.get('original_filename', 'donnees') or 'donnees').rsplit('.', 1)[0]
    return send_file(buf, as_attachment=True,
                      download_name=f'{fname}_questionnaire.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/data/reset')
def reset_data():
    for key in ['data_path', 'child_path']:
        path = session.get(key)
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
    session.clear()
    flash('Données effacées.', 'info')
    return redirect(url_for('index'))


@app.route('/data/raw')
def raw_data():
    df = get_dataframe()
    if df is None:
        return redirect(url_for('upload'))

    overview   = overview_stats(df)
    chart_miss = missing_bar_chart(df)
    chart_heat = missing_heatmap(df)
    chart_corr = correlation_matrix(df)
    
    # Analyse systématique complète
    sys_analysis = systematic_analysis(df)
    chart_quality = data_quality_gauge(sys_analysis['quality'])
    chart_distrib = distributions_chart(df)
    chart_composition = composition_chart(sys_analysis)
    
    # Analyse descriptive avec tableaux et graphiques
    desc_analysis = descriptive_summary(df)

    # Preview table (first 100 rows, paginated client-side via DataTables)
    preview_html = df.head(200).to_html(
        classes='table table-sm table-hover table-bordered mb-0',
        border=0, index=True, na_rep='—', table_id='rawTable')

    return render_template('raw_data.html',
                           overview=overview,
                           preview_html=preview_html,
                           chart_miss=chart_miss,
                           chart_heat=chart_heat,
                           chart_corr=chart_corr,
                           chart_quality=chart_quality,
                           chart_distrib=chart_distrib,
                           chart_composition=chart_composition,
                           sys_analysis=sys_analysis,
                           desc_analysis=desc_analysis,
                           meta=session.get('data_meta', {}))


@app.route('/map')
def map_view():
    """Carte géographique interactive — points GPS avec infobulles."""
    df = get_dataframe()
    if df is None:
        flash("Veuillez d'abord importer un fichier.", 'warning')
        return redirect(url_for('upload'))

    import re as _re, ast as _ast

    # ── 1. Parseur GPS ──────────────────────────────────────────────────
    def _parse_gps(val):
        """
        Accepte :
          "5.18 -4.60 29.0 4.4"   → (5.18, -4.60)
          "[5.18, -4.60]"          → (5.18, -4.60)
          "5.18,-4.60"             → (5.18, -4.60)
        """
        if not val or (isinstance(val, float) and pd.isna(val)):
            return None, None
        s = str(val).strip()
        # Format KoboToolbox "lat lon alt prec"
        parts = s.replace(',', ' ').split()
        if len(parts) >= 2:
            try:
                lat, lon = float(parts[0]), float(parts[1])
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    return round(lat, 6), round(lon, 6)
            except ValueError:
                pass
        # Format "[lat, lon]"
        m = _re.search(r'\[?\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\]?', s)
        if m:
            try:
                lat, lon = float(m.group(1)), float(m.group(2))
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    return round(lat, 6), round(lon, 6)
            except ValueError:
                pass
        return None, None

    # ── 2. Colonnes GPS candidates (par priorité) ───────────────────────
    # On accepte deux dispositions :
    #   (a) colonne unique combinée (Kobo : "_geolocation", "gps"…)
    #   (b) deux colonnes séparées lat/lon (Excel "à la main")
    GPS_PRIORITY = ['_geolocation', 'start_gps', 'end_gps',
                    'gps', 'location', 'position', 'coordonnees']
    gps_col = None
    lat_col = None
    lon_col = None

    for cand in GPS_PRIORITY:
        if cand in df.columns:
            gps_col = cand
            break

    if gps_col is None:
        # Tentative : colonnes lat / lon séparées.
        # On accepte que "latitude" / "longitude" apparaissent en fin ou
        # entourés de [_ ] dans le nom (ex: '_Géolocalisation_latitude').
        LAT_PATTERNS = [
            r'(?:^|[_ ])latitude(?:[_ ]|$)',
            r'(?:^|[_ ])lat(?:[_ ]|$)',
            r'^_?y$',
            r'coord[_ ]y',
        ]
        LON_PATTERNS = [
            r'(?:^|[_ ])longitude(?:[_ ]|$)',
            r'(?:^|[_ ])lon(?:[_ ]|$)',
            r'(?:^|[_ ])lng(?:[_ ]|$)',
            r'(?:^|[_ ])long(?:[_ ]|$)',
            r'^_?x$',
            r'coord[_ ]x',
        ]
        for pat in LAT_PATTERNS:
            for col in df.columns:
                if _re.search(pat, col, _re.IGNORECASE):
                    lat_col = col
                    break
            if lat_col:
                break
        for pat in LON_PATTERNS:
            for col in df.columns:
                if _re.search(pat, col, _re.IGNORECASE) and col != lat_col:
                    lon_col = col
                    break
            if lon_col:
                break

        # Sinon : nom de colonne avec un mot-clé GPS au sens strict.
        # Word boundaries pour éviter de matcher "geo" dans "Rougeole".
        if not (lat_col and lon_col):
            STRICT_GPS_PAT = _re.compile(
                r'(?:^|[\s_\-/])(gps|geolocation|geo[_\-]?point|geo[_\-]?code|'
                r'coord|coordonn[eé]es?|location|position|gps_?point)(?:$|[\s_\-/])',
                _re.IGNORECASE)
            for col in df.columns:
                if STRICT_GPS_PAT.search(col):
                    gps_col = col
                    break

        # Dernier recours : détection par les VALEURS — on échantillonne chaque
        # colonne objet et on garde celle dont le plus de cellules parsent comme GPS.
        if gps_col is None and not (lat_col and lon_col):
            best_col, best_hits = None, 0
            sample = df.head(min(20, len(df)))
            for col in df.columns:
                if not (df[col].dtype == object or pd.api.types.is_string_dtype(df[col])):
                    continue
                hits = 0
                for v in sample[col]:
                    la, lo = _parse_gps(v)
                    if la is not None and lo is not None:
                        hits += 1
                if hits > best_hits and hits >= 2:   # au moins 2 hits pour être convaincant
                    best_col, best_hits = col, hits
            if best_col:
                gps_col = best_col

    # ── 3. Colonnes contexte — détection flexible avec normalisation ────
    # Normalise pour absorber accents et casse : "Région sanitaire" → "region sanitaire"
    import unicodedata as _ud
    def _norm(s):
        s = str(s)
        # NFKD décompose les caractères accentués ; on supprime les marques combinantes
        s = ''.join(c for c in _ud.normalize('NFKD', s) if not _ud.combining(c))
        return s.lower()

    def _find_col(patterns, cols):
        """Retourne la 1re colonne dont le nom (normalisé) matche un motif."""
        normed = [(_norm(c), c) for c in cols]
        for pat in patterns:
            rx = _re.compile(pat, _re.IGNORECASE)
            for n, original in normed:
                if rx.search(n):
                    return original
        return None

    cols = df.columns.tolist()
    ctx = {
        # Région : "Région sanitaire", "Nom de la région", "Province", "Region 1"
        'region':      _find_col([
            r'region[_\s\-]?sanit',
            r'\bregion\b',
            r'\bprovince\b',
            r'zone[_\s\-]?admin',
        ], cols),
        # District : "District sanitaire", "Nom du district", "Département"
        'district':    _find_col([
            r'district[_\s\-]?sanit',
            r'\bdistrict\b',
            r'\bdepartement\b',
            r'sous[_\s\-]?(region|prefecture)',
        ], cols),
        # Aire de santé : "Aire de santé", "Aire sanitaire", "Secteur", "Localité"
        'aire':        _find_col([
            r'aire[_\s\-]?(de[_\s\-]?)?sant',
            r'aire[_\s\-]?sanit',
            r'\baire\b',
            r'\bsecteur\b',
            r'\blocalite\b',
            r'\bvillage\b',
        ], cols),
        # Superviseur : "Nom du superviseur", "Superviseur de l'enquête"
        'superviseur': _find_col([
            r'\bsupervis(eur|euse|ion)?\b',
            r'supervis',                    # filet de sécurité
        ], cols),
        # Enquêteur : "Nom de l'enquêteur", "Agent de collecte", "Collecteur"
        'enqueteur':   _find_col([
            r'\benqueteur\b',
            r'\benquetrice\b',
            r'agent[_\s\-]?(de[_\s\-]?)?collect',
            r'\bcollecteur\b',
            r'\benquet',                    # filet de sécurité
        ], cols),
        # Date : "Date de collecte", "Date enquête", "Date interview"
        'date':        _find_col([
            r'date[_\s\-]?(de[_\s\-]?)?collect',
            r'date[_\s\-]?(de[_\s\-]?)?(enquet|interview)',
            r'^date\b',
        ], cols),
        # Zone : "Zone d'habitation", "Zone de résidence", "Milieu", "Urbain/rural"
        'zone':        _find_col([
            r'zone[_\s\-]?(d[\'_\s\-])?(habit|resid)',
            r'zone[_\s\-]?(rural|urbain)',
            r'\bzone\b',
            r'\bmilieu\b',
            r'\burbain\b',
            r'\brural\b',
        ], cols),
    }

    # ── Durée d'enquête — calcul automatique depuis start/end ODK ──────
    # KoboToolbox stocke l'heure de début/fin dans les colonnes 'start' et 'end'
    # (questions méta ODK). Si présentes, on calcule la durée réelle d'entretien.
    start_col = _find_col([r'^start$', r'^heure_debut$', r'^debut$', r'^start_time$'], cols)
    end_col   = _find_col([r'^end$',   r'^heure_fin$',   r'^fin$',   r'^end_time$'],   cols)
    # Colonne de durée explicite (si le formulaire la calcule)
    explicit_dur_col = _find_col(
        [r'duree_entretien', r'duree_enquete', r'temps_enquete',
         r'interview_duration', r'survey_duration', r'duree_interview'],
        cols)

    duration_mode = 'none'   # 'computed', 'explicit', 'none'
    if start_col and end_col:
        duration_mode = 'computed'
    elif explicit_dur_col:
        duration_mode = 'explicit'

    # ── 4. Construction de geo_data ─────────────────────────────────────
    geo_data   = []
    no_gps_cnt = 0
    out_of_range_cnt = 0
    parse_fail_cnt = 0

    def _to_float(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        try:
            return float(str(v).replace(',', '.').strip())
        except (ValueError, TypeError):
            return None

    for idx, row in df.iterrows():
        lat, lon = None, None
        if gps_col:
            lat, lon = _parse_gps(row[gps_col])
            if lat is None or lon is None:
                # cellule présente mais format pas reconnu vs cellule vide
                v = row[gps_col]
                if v is not None and not (isinstance(v, float) and pd.isna(v)):
                    parse_fail_cnt += 1
        elif lat_col and lon_col:
            la = _to_float(row.get(lat_col))
            lo = _to_float(row.get(lon_col))
            if la is not None and lo is not None:
                if -90 <= la <= 90 and -180 <= lo <= 180:
                    lat, lon = round(la, 6), round(lo, 6)
                else:
                    out_of_range_cnt += 1
        if lat is None or lon is None:
            no_gps_cnt += 1
            continue

        def _v(key):
            col = ctx.get(key)
            if col and col in row.index:
                v = row[col]
                return str(v) if pd.notna(v) else '—'
            return '—'

        # ── Calcul de la durée d'enquête ────────────────────────────────
        if duration_mode == 'computed':
            try:
                t_start = pd.to_datetime(row[start_col])
                t_end   = pd.to_datetime(row[end_col])
                diff_s  = (t_end - t_start).total_seconds()
                if diff_s > 0:
                    mins = int(diff_s // 60)
                    secs = int(diff_s % 60)
                    duration_str = f'{mins} min {secs:02d} s' if mins < 60 \
                                   else f'{mins // 60} h {mins % 60:02d} min'
                else:
                    duration_str = '—'
            except Exception:
                duration_str = '—'
        elif duration_mode == 'explicit':
            v = row.get(explicit_dur_col)
            duration_str = f'{v} min' if pd.notna(v) else '—'
        else:
            duration_str = None   # None = champ absent du formulaire

        geo_data.append({
            'lat':          lat,
            'lon':          lon,
            'region':       _v('region'),
            'district':     _v('district'),
            'aire':         _v('aire'),
            'superviseur':  _v('superviseur'),
            'enqueteur':    _v('enqueteur'),
            'date':         _v('date'),
            'duration':     duration_str,   # None si le form ne le capture pas
            'zone':         _v('zone'),
            'idx':          int(idx) + 1,
        })

    # ── 5. Centre & zoom ────────────────────────────────────────────────
    if geo_data:
        lats = [p['lat'] for p in geo_data]
        lons = [p['lon'] for p in geo_data]
        center_lat = round(sum(lats) / len(lats), 5)
        center_lon = round(sum(lons) / len(lons), 5)
        lat_range  = max(lats) - min(lats)
        lon_range  = max(lons) - min(lons)
        spread     = max(lat_range, lon_range)
        zoom = 14 if spread < 0.01 else 12 if spread < 0.05 else \
               10 if spread < 0.2  else 8  if spread < 1    else \
               6  if spread < 5    else 5
    else:
        center_lat, center_lon, zoom = 5.35, -4.00, 7

    # ── 6. Stats par région ─────────────────────────────────────────────
    region_counts = {}
    for p in geo_data:
        r = p['region']
        region_counts[r] = region_counts.get(r, 0) + 1

    # Couleurs par région (palette lisible)
    PALETTE = ['#2E86C1','#17A589','#E67E22','#8E44AD','#C0392B',
               '#1A5276','#148F77','#B7770D','#6C3483','#922B21',
               '#117A65','#A04000','#154360','#7D6608','#4D5656']
    regions_sorted = sorted(region_counts.keys())
    region_colors  = {r: PALETTE[i % len(PALETTE)]
                      for i, r in enumerate(regions_sorted)}
    for p in geo_data:
        p['color'] = region_colors.get(p['region'], '#1A5276')

    # Libellés des colonnes détectées (pour le panel d'info)
    ctx_labels = {k: v for k, v in ctx.items() if v}
    if duration_mode == 'computed':
        ctx_labels['durée enquête'] = f'{start_col} → {end_col} (calculée)'
    elif duration_mode == 'explicit':
        ctx_labels['durée enquête'] = explicit_dur_col

    # ── Diagnostic : si rien de trouvé, on liste ce qui a été inspecté ──
    diagnostic = None
    if not geo_data:
        sample_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c]) or df[c].dtype == object][:30]
        if not gps_col and not (lat_col and lon_col):
            reason = "Aucune colonne GPS détectée. Renommez une colonne en 'latitude'/'longitude', '_geolocation' ou 'gps'."
        elif parse_fail_cnt:
            reason = f"Colonne '{gps_col}' trouvée mais le format des cellules n'est pas reconnu (échec de parsing : {parse_fail_cnt})."
        elif out_of_range_cnt:
            reason = f"Coordonnées hors limites : lat ∉ [-90,90] ou lon ∉ [-180,180] ({out_of_range_cnt} ligne(s))."
        else:
            reason = "Toutes les cellules GPS sont vides."
        diagnostic = {
            'reason': reason,
            'gps_col': gps_col,
            'lat_col': lat_col,
            'lon_col': lon_col,
            'sample_cols': sample_cols,
            'parse_fail_cnt': parse_fail_cnt,
            'out_of_range_cnt': out_of_range_cnt,
            'no_gps_cnt': no_gps_cnt,
            'total_rows': int(len(df)),
        }

    label = gps_col or (f'{lat_col} / {lon_col}' if lat_col and lon_col else '—')
    return render_template(
        'map_view.html',
        geo_data      = geo_data,
        no_gps_cnt    = no_gps_cnt,
        center_lat    = center_lat,
        center_lon    = center_lon,
        zoom          = zoom,
        region_counts = region_counts,
        region_colors = region_colors,
        ctx_labels    = ctx_labels,
        gps_col       = label,
        diagnostic    = diagnostic,
        duration_mode = duration_mode,
        meta          = session.get('data_meta', {}),
        total_pts     = len(geo_data),
    )


@app.route('/dashboard')
def dashboard():
    """Tableau de bord synthétique."""
    df = get_dataframe()
    if df is None:
        flash('Veuillez d\'abord importer un fichier.', 'warning')
        return redirect(url_for('upload'))
    meta = session.get('data_meta', {})
    return render_template('dashboard.html', meta=meta)


# ─── DESCRIPTIVE ──────────────────────────────────────────────────────────────

@app.route('/descriptive', methods=['GET', 'POST'])
def descriptive():
    df = get_dataframe()
    if df is None:
        return redirect(url_for('upload'))
    var_types = get_var_types(df)
    cat_vars  = [v for v, t in var_types.items() if t in ('categorielle',)]
    num_vars  = [v for v, t in var_types.items() if t == 'continue']
    bin_groups = detect_binary_groups(df)

    results_cat, results_num, results_bin = [], [], []
    selected_cat, selected_num, selected_bin_groups = [], [], []

    if request.method == 'POST':
        selected_cat = request.form.getlist('cat_vars')
        selected_num = request.form.getlist('num_vars')
        selected_bin_groups = request.form.getlist('bin_groups')
        # Persiste les sélections pour les retrouver au prochain GET
        session['descriptive_form'] = {
            'cat': selected_cat,
            'num': selected_num,
            'bin': selected_bin_groups,
        }
    else:
        saved = session.get('descriptive_form')
        if saved:
            selected_cat = saved.get('cat', [])
            selected_num = saved.get('num', [])
            selected_bin_groups = saved.get('bin', [])
        else:
            # Première visite (pas encore de sélection enregistrée) : calcule
            # automatiquement sur toutes les variables plutôt que de laisser
            # une page vide en attente d'une action manuelle — les stats
            # descriptives par variable ne demandent aucun jugement quant au
            # choix des variables, contrairement à un croisement ciblé.
            selected_cat = cat_vars
            selected_num = num_vars
            selected_bin_groups = [g['parent'] for g in bin_groups]

    # Recalcule à partir des sélections (qu'elles viennent du POST ou de la session)
    for var in selected_cat:
        if var in df.columns:
            r = tris_a_plat(df, var)
            r['auto_comment'] = auto_comment_categorical(r)
            r['user_comment'] = _get_user_comment('descriptive_cat', var)
            results_cat.append(r)

    for var in selected_num:
        if var in df.columns:
            r = statistiques_continues(df, var)
            r['auto_comment'] = auto_comment_continuous(r)
            r['user_comment'] = _get_user_comment('descriptive_num', var)
            results_num.append(r)

    for parent in selected_bin_groups:
        group = next((g for g in bin_groups if g['parent'] == parent), None)
        if group:
            r = {'parent': parent, **analyse_binaire_groupe(df, group['vars'])}
            r['auto_comment'] = auto_comment_binary_group(r)
            r['user_comment'] = _get_user_comment('descriptive_bin', parent)
            results_bin.append(r)

    return render_template('descriptive.html',
                           cat_vars=cat_vars, num_vars=num_vars,
                           bin_groups=bin_groups,
                           results_cat=results_cat, results_num=results_num,
                           results_bin=results_bin,
                           selected_cat=selected_cat,
                           selected_num=selected_num,
                           selected_bin_groups=selected_bin_groups)


# ─── CROSSTABS ────────────────────────────────────────────────────────────────

@app.route('/crosstabs', methods=['GET', 'POST'])
def crosstabs():
    """Analyse Croisée Dynamique (style tableau croisé dynamique Excel)."""
    df = get_dataframe()
    if df is None:
        return redirect(url_for('upload'))
    var_types = get_var_types(df)
    cat_vars = [v for v, t in var_types.items() if t in ('categorielle', 'binaire')]
    num_vars = [v for v, t in var_types.items() if t == 'continue']
    all_vars = list(df.columns)
    result = None
    rows = []
    cols = []
    value_var = None
    agg = 'count'
    pct_type = 'none'
    filters = {}

    if request.method == 'POST':
        rows = [r for r in request.form.getlist('rows') if r]
        cols = [c for c in request.form.getlist('cols') if c]
        value_var = (request.form.get('value_var') or '').strip() or None
        agg = request.form.get('agg', 'count')
        pct_type = request.form.get('pct_type', 'none')
        # Filtres : 'filter_<var>' = liste de valeurs
        for key in request.form.keys():
            if key.startswith('filter_'):
                var = key[len('filter_'):]
                vals = [v for v in request.form.getlist(key) if v]
                if vals:
                    filters[var] = vals
        session['crosstabs_form'] = {
            'rows': rows, 'cols': cols, 'value_var': value_var,
            'agg': agg, 'pct_type': pct_type, 'filters': filters,
        }
    else:
        saved = session.get('crosstabs_form') or {}
        rows     = [r for r in saved.get('rows', []) if r in df.columns]
        cols     = [c for c in saved.get('cols', []) if c in df.columns]
        value_var = saved.get('value_var')
        agg      = saved.get('agg', 'count')
        pct_type = saved.get('pct_type', 'none')
        filters  = saved.get('filters', {}) or {}

    if rows:
        try:
            result = tableau_croise_dynamique(
                df, rows=rows, cols=cols,
                value_var=value_var, agg=agg,
                pct_type=pct_type, filters=filters,
            )
            # Commentaire automatique seulement si comptage 2D simple
            if agg == 'count' and len(rows) == 1 and len(cols) == 1:
                # Construire un objet compatible avec auto_comment_crosstab
                legacy = {
                    'row_var': rows[0], 'col_var': cols[0],
                    'n': result['n'], 'chi2': result['chi2'],
                    'p_val': result['p_val'], 'dof': result['dof'],
                    'cramers_v': result['cramers_v'],
                    'sig_label': result['sig_label'],
                    'assoc_label': result['assoc_label'],
                    'pct_label': result['pct_label'] or '% ligne',
                }
                try:
                    result['auto_comment'] = auto_comment_crosstab(legacy)
                except Exception:
                    result['auto_comment'] = ''
            else:
                result['auto_comment'] = (
                    f"Tableau croisé dynamique : {agg} sur "
                    f"{value_var or 'observations'} — "
                    f"{' × '.join(rows)} vs {' × '.join(cols) if cols else '(aucune colonne)'}"
                )
            key = '|'.join(rows) + '||' + '|'.join(cols)
            result['user_comment'] = _get_user_comment('crosstabs', key)
        except Exception as e:
            if request.method == 'POST':
                flash(f'Erreur : {str(e)}', 'danger')

    # Modalités disponibles pour le panneau de filtres (≤ 30 valeurs)
    filter_options = {}
    for v in cat_vars:
        try:
            vals = [str(x) for x in df[v].dropna().unique()[:30]]
            if vals:
                filter_options[v] = vals
        except Exception:
            pass

    return render_template('crosstabs.html',
                            cat_vars=cat_vars, num_vars=num_vars,
                            all_vars=all_vars,
                            filter_options=filter_options,
                            result=result, rows=rows, cols=cols,
                            value_var=value_var, agg=agg,
                            pct_type=pct_type, filters=filters)


# ─── MULTI-ENQUÊTE ────────────────────────────────────────────────────────────

def _load_multi_surveys() -> dict:
    """Charge les enquêtes enregistrées en session sous forme {nom: DataFrame}."""
    surveys = {}
    entries = session.get('multi_surveys', []) or []
    for entry in entries:
        path = entry.get('path')
        name = entry.get('name') or os.path.basename(path or '')
        if path and os.path.exists(path):
            try:
                surveys[name] = pd.read_excel(path)
            except Exception:
                pass
    return surveys


def _compute_ms_comparison(merged, variable):
    """Calcule la comparaison d'UNE variable entre enquêtes. Renvoie un dict
    {kind: 'continuous'|'categorical', ...} ou {error: '...'}."""
    try:
        if variable not in merged.columns:
            return {'variable': variable, 'error': f"Variable '{variable}' absente après alignement."}
        serie = pd.to_numeric(merged[variable], errors='coerce')
        ratio_num = serie.notna().mean()
        if ratio_num > 0.8 and serie.nunique() > 10:
            r = compare_continuous_by_survey(merged, variable)
            r['kind'] = 'continuous'
        else:
            r = compare_categorical_by_survey(merged, variable)
            r['kind'] = 'categorical'
        r['variable'] = variable
        return r
    except Exception as e:
        return {'variable': variable, 'error': str(e)}


def _ms_build_xlsx(mv_result: dict, mv_method: str) -> bytes:
    """Construit un fichier XLSX à partir du résultat d'une analyse multivariée."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='1a6fa8')
    title_font   = Font(bold=True, size=12)

    def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
        ws.cell(row=start_row, column=1, value=df.index.name or 'Index')
        ws.cell(row=start_row, column=1).font = header_font
        ws.cell(row=start_row, column=1).fill = header_fill
        for ci, col in enumerate(df.columns, start=2):
            c = ws.cell(row=start_row, column=ci, value=str(col))
            c.font = header_font
            c.fill = header_fill
        for ri, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
            ws.cell(row=ri, column=1, value=str(idx))
            for ci, val in enumerate(row, start=2):
                ws.cell(row=ri, column=ci, value=round(float(val), 4) if isinstance(val, (float, int)) else str(val))
        for col_cells in ws.columns:
            length = max(len(str(c.value or '')) for c in col_cells) + 2
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length, 40)

    def _write_html_table(ws, html: str, title: str, start_row: int = 1) -> int:
        try:
            dfs = pd.read_html(html)
            if dfs:
                df = dfs[0]
                t = ws.cell(row=start_row, column=1, value=title)
                t.font = title_font
                start_row += 1
                for ci, col in enumerate(df.columns, start=1):
                    c = ws.cell(row=start_row, column=ci, value=str(col))
                    c.font = header_font
                    c.fill = header_fill
                start_row += 1
                for _, row in df.iterrows():
                    for ci, val in enumerate(row, start=1):
                        try:
                            ws.cell(row=start_row, column=ci, value=float(val))
                        except Exception:
                            ws.cell(row=start_row, column=ci, value=str(val) if pd.notna(val) else '')
                    start_row += 1
                return start_row + 1
        except Exception:
            pass
        return start_row

    method_label = {
        'pca': 'ACP', 'mca': 'ACM', 'ca': 'AFC', 'clustering': 'Classification K-Means'
    }.get(mv_method, mv_method)

    # Feuille synthèse
    ws_info = wb.create_sheet('Synthèse')
    ws_info['A1'] = f'Méthode : {method_label}'
    ws_info['A1'].font = title_font
    ws_info['A2'] = f'Observations : {mv_result.get("n_obs", "—")}'
    ws_info['A3'] = f'Variables : {mv_result.get("n_vars", mv_result.get("n_vars", "—"))}'
    if 'variables' in mv_result:
        ws_info['A5'] = 'Variables utilisées :'
        ws_info['A5'].font = Font(bold=True)
        for i, v in enumerate(mv_result['variables'], start=6):
            ws_info.cell(row=i, column=1, value=v)

    row = 1
    if mv_method == 'pca':
        ws_eig = wb.create_sheet('Valeurs propres')
        row = _write_html_table(ws_eig, mv_result.get('eig_table', ''), 'Tableau des valeurs propres', 1)
        ws_load = wb.create_sheet('Saturations (loadings)')
        _write_html_table(ws_load, mv_result.get('loadings_table', ''), 'Saturations factorielles', 1)

    elif mv_method == 'mca':
        ws_eig = wb.create_sheet('Valeurs propres (ACM)')
        _write_html_table(ws_eig, mv_result.get('eig_table', ''), 'Tableau des axes', 1)

    elif mv_method == 'ca':
        ws_ct = wb.create_sheet('Tableau de contingence')
        _write_html_table(ws_ct, mv_result.get('ct_html', ''), 'Tableau de contingence', 1)
        ws_eig = wb.create_sheet('Valeurs propres (AFC)')
        _write_html_table(ws_eig, mv_result.get('eig_table', ''), 'Tableau des axes', 1)

    elif mv_method == 'clustering':
        ws_eff = wb.create_sheet('Effectifs des classes')
        _write_html_table(ws_eff, mv_result.get('counts_table', ''), 'Effectifs par classe', 1)
        ws_moy = wb.create_sheet('Moyennes par classe')
        _write_html_table(ws_moy, mv_result.get('means_table', ''), 'Moyennes par classe', 1)
        ws_info['A7'] = f'Indice de silhouette : {mv_result.get("silhouette", "—")}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.route('/multi-survey', methods=['GET', 'POST'])
def multi_survey():
    """Analyse multi-enquête avec alignement DPF/PHAKTS — comparaisons multiples persistées."""
    surveys_meta = session.get('multi_surveys', []) or []
    auto_mode = session.get('ms_auto_mode', False)
    mode = request.form.get('align_mode') or session.get('ms_mode', 'phakts')
    compared_vars = list(session.get('ms_compared_vars', []) or [])
    results = []
    auto_results = []
    coverage_html = None
    merged_meta = None
    common_vars = []

    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'upload':
            files = request.files.getlist('files')
            added = 0
            for f in files:
                if not f or not f.filename or not allowed_file(f.filename):
                    continue
                fname = str(uuid.uuid4()) + '_' + secure_filename(f.filename)
                raw_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ms_' + fname)
                f.save(raw_path)
                try:
                    df_raw = load_excel(raw_path)
                    df_clean = clean_kobo_dataframe(df_raw)
                    cleaned_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ms_clean_' + fname)
                    df_clean.to_excel(cleaned_path, index=False, engine='openpyxl')
                    surveys_meta.append({
                        'name': f.filename,
                        'path': cleaned_path,
                        'n_obs': int(len(df_clean)),
                        'n_vars': int(df_clean.shape[1]),
                        'phakts_vars': int(sum(1 for c in df_clean.columns if is_phakts_coded(c))),
                    })
                    added += 1
                except Exception as e:
                    flash(f"Erreur sur « {f.filename} » : {e}", 'danger')
                finally:
                    try: os.remove(raw_path)
                    except Exception: pass
            session['multi_surveys'] = surveys_meta
            if added:
                flash(f"{added} enquête(s) ajoutée(s) — total : {len(surveys_meta)}", 'success')

        elif action == 'remove':
            idx = int(request.form.get('idx', '-1'))
            if 0 <= idx < len(surveys_meta):
                p = surveys_meta[idx].get('path')
                try: os.remove(p)
                except Exception: pass
                del surveys_meta[idx]
                session['multi_surveys'] = surveys_meta
                flash("Enquête retirée.", 'info')

        elif action == 'reset':
            for entry in surveys_meta:
                try: os.remove(entry.get('path'))
                except Exception: pass
            session['multi_surveys'] = []
            surveys_meta = []
            session['ms_compared_vars'] = []
            compared_vars = []
            flash("Toutes les enquêtes ont été effacées.", 'info')

        elif action == 'add_compare':
            # Ajoute une ou plusieurs variables à comparer
            new_vars = [v for v in request.form.getlist('variables') if v]
            session['ms_mode'] = mode
            added = 0
            for v in new_vars:
                if v not in compared_vars:
                    compared_vars.append(v)
                    added += 1
            session['ms_compared_vars'] = compared_vars
            if added:
                flash(f"{added} variable(s) ajoutée(s) aux comparaisons (total : {len(compared_vars)}).", 'success')
            elif new_vars:
                flash("Variable(s) déjà présente(s) dans les comparaisons.", 'info')

        elif action == 'remove_compare':
            v = request.form.get('variable', '')
            if v in compared_vars:
                compared_vars.remove(v)
                session['ms_compared_vars'] = compared_vars
                flash(f"« {v} » retirée des comparaisons.", 'info')

        elif action == 'clear_compare':
            session['ms_compared_vars'] = []
            compared_vars = []
            flash("Toutes les comparaisons ont été retirées.", 'info')

        elif action == 'change_mode':
            session['ms_mode'] = mode

        elif action == 'toggle_auto':
            auto_mode = request.form.get('auto_mode') == 'on'
            session['ms_auto_mode'] = auto_mode

        elif action == 'kobo_refresh_ms':
            token = session.get('kobo_token')
            instance = session.get('kobo_instance')
            if not token:
                flash("Connectez-vous d'abord à KoboToolbox (menu Connexion KoboToolbox).", 'warning')
            else:
                res = list_assets(token, instance=instance or None)
                if res.get('success'):
                    session['ms_kobo_assets'] = res['assets']
                    session['kobo_instance'] = res.get('instance', instance)
                    flash(f"{res['total']} formulaire(s) KoboToolbox chargé(s).", 'success')
                else:
                    flash(f"KoboToolbox : {res.get('error', 'Erreur inconnue')}", 'danger')

        elif action == 'kobo_add':
            token = session.get('kobo_token')
            instance = session.get('kobo_instance')
            uid = request.form.get('kobo_uid', '').strip()
            if not token:
                flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
            elif not uid:
                flash("Sélectionnez un formulaire KoboToolbox.", 'warning')
            else:
                res = kobo_load_data(token, uid, instance=instance or None)
                if res.get('success'):
                    df_kobo = res['df']
                    asset_name = request.form.get('kobo_asset_name') or f'Kobo_{uid[:8]}'
                    fname = f"ms_kobo_{uuid.uuid4().hex[:8]}.xlsx"
                    fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                    df_kobo.to_excel(fpath, index=False, engine='openpyxl')
                    surveys_meta.append({
                        'name': asset_name,
                        'path': fpath,
                        'n_obs': int(len(df_kobo)),
                        'n_vars': int(df_kobo.shape[1]),
                        'phakts_vars': int(sum(1 for c in df_kobo.columns if is_phakts_coded(c))),
                        'source': 'kobo',
                    })
                    session['multi_surveys'] = surveys_meta
                    flash(f"Enquête KoboToolbox « {asset_name} » ajoutée ({len(df_kobo)} obs.).", 'success')
                else:
                    flash(f"Chargement KoboToolbox : {res.get('error', 'Erreur')}", 'danger')

        elif action == 'run_mv_ms':
            session['ms_mv_params'] = {
                'method':       request.form.get('mv_method', 'pca'),
                'variables':    request.form.getlist('mv_variables'),
                'n_components': int(request.form.get('mv_n_components', 2)),
                'n_clusters':   int(request.form.get('mv_n_clusters', 3)),
                'ca_row':       request.form.get('mv_ca_row', ''),
                'ca_col':       request.form.get('mv_ca_col', ''),
            }

        elif action == 'clear_mv_ms':
            session.pop('ms_mv_params', None)

    surveys = _load_multi_surveys()

    if len(surveys) >= 1:
        try:
            cov = variable_coverage(surveys, mode=mode)
            coverage_html = cov.head(80).to_html(
                classes='table table-sm table-bordered text-center',
                border=0, na_rep='—')
        except Exception:
            pass
        try:
            common_vars = common_variables(surveys, mode=mode)
        except Exception:
            common_vars = []

        if len(surveys) >= 2:
            try:
                merged = merge_surveys(surveys, mode=mode)
                merged_meta = {
                    'n_obs': int(len(merged)),
                    'n_vars': int(merged.shape[1] - 1),
                    'n_surveys': len(surveys),
                }
            except Exception as e:
                merged = None
                flash(f"Erreur d'analyse : {e}", 'danger')

            if auto_mode and merged is not None:
                # Interopérabilité : indicateurs communs détectés automatiquement
                # via le xType PHAKTS porté par le nom de chaque colonne (voir
                # modules/multi_survey.py::auto_analyze) — aucune sélection
                # manuelle de variable nécessaire.
                try:
                    auto_results = auto_analyze(surveys, mode=mode)
                except Exception as e:
                    auto_results = []
                    flash(f"Erreur d'analyse automatique : {e}", 'danger')
            elif compared_vars and merged is not None:
                for v in compared_vars:
                    results.append(_compute_ms_comparison(merged, v))

    # Variables proposées (union + variables PHAKTS)
    candidate_vars = []
    if surveys:
        all_cols_by_survey = []
        for df in surveys.values():
            if mode == 'phakts':
                all_cols_by_survey.append(set(phakts_radical(c) for c in df.columns))
            else:
                all_cols_by_survey.append(set(df.columns))
        if all_cols_by_survey:
            union = sorted(set().union(*all_cols_by_survey) - {'__survey__'})
            candidate_vars = union

    # ── Analyse multivariée sur jeu fusionné ─────────────────────────────────
    mv_result = None
    mv_method = None
    mv_params = session.get('ms_mv_params') or {}
    mv_num_vars = []
    mv_cat_vars = []
    mv_merged_df = None

    if surveys:
        try:
            mv_merged_df = merge_surveys(surveys, mode=mode)
            vt = get_var_types(mv_merged_df)
            mv_num_vars = [v for v, t in vt.items() if t in ('continue', 'binaire') and v != '__survey__']
            mv_cat_vars = [v for v, t in vt.items() if t == 'categorielle' and v != '__survey__']
        except Exception:
            pass

    if mv_params and mv_merged_df is not None:
        mv_method = mv_params.get('method', 'pca')
        mv_sel    = [v for v in mv_params.get('variables', []) if v in (mv_merged_df.columns.tolist())]
        mv_nc     = mv_params.get('n_components', 2)
        mv_nk     = mv_params.get('n_clusters', 3)
        mv_row    = mv_params.get('ca_row', '')
        mv_col    = mv_params.get('ca_col', '')
        try:
            if mv_method == 'pca' and len(mv_sel) >= 2:
                mv_result = run_pca(mv_merged_df, mv_sel, mv_nc)
            elif mv_method == 'mca' and len(mv_sel) >= 2:
                mv_result = run_mca(mv_merged_df, mv_sel, mv_nc)
            elif mv_method == 'ca' and mv_row and mv_col:
                mv_result = run_ca(mv_merged_df, mv_row, mv_col)
            elif mv_method == 'clustering' and len(mv_sel) >= 2:
                mv_result = run_clustering(mv_merged_df, mv_sel, mv_nk)
        except Exception as e:
            flash(f"Analyse multivariée : {e}", 'danger')

    return render_template('multi_survey.html',
                            surveys_meta=surveys_meta,
                            mode=mode,
                            coverage_html=coverage_html,
                            common_vars=common_vars,
                            candidate_vars=candidate_vars,
                            compared_vars=compared_vars,
                            results=results,
                            auto_mode=auto_mode,
                            auto_results=auto_results,
                            merged_meta=merged_meta,
                            kobo_connected=bool(session.get('kobo_token')),
                            ms_kobo_assets=session.get('ms_kobo_assets', []),
                            mv_result=mv_result,
                            mv_method=mv_method,
                            mv_params=mv_params,
                            mv_num_vars=mv_num_vars,
                            mv_cat_vars=mv_cat_vars)


@app.route('/multi-survey/download-xlsx')
def multi_survey_download_xlsx():
    """Télécharge en XLSX les résultats de l'analyse multivariée multi-enquête."""
    mv_params = session.get('ms_mv_params') or {}
    if not mv_params:
        flash("Lancez d'abord une analyse multivariée.", 'warning')
        return redirect(url_for('multi_survey'))

    surveys = _load_multi_surveys()
    if not surveys:
        flash("Aucune enquête chargée.", 'warning')
        return redirect(url_for('multi_survey'))

    mode = session.get('ms_mode', 'phakts')
    try:
        merged = merge_surveys(surveys, mode=mode)
    except Exception as e:
        flash(f"Fusion impossible : {e}", 'danger')
        return redirect(url_for('multi_survey'))

    mv_method = mv_params.get('method', 'pca')
    mv_sel    = [v for v in mv_params.get('variables', []) if v in merged.columns.tolist()]
    mv_nc     = mv_params.get('n_components', 2)
    mv_nk     = mv_params.get('n_clusters', 3)
    mv_row    = mv_params.get('ca_row', '')
    mv_col    = mv_params.get('ca_col', '')

    try:
        if mv_method == 'pca' and len(mv_sel) >= 2:
            result = run_pca(merged, mv_sel, mv_nc)
        elif mv_method == 'mca' and len(mv_sel) >= 2:
            result = run_mca(merged, mv_sel, mv_nc)
        elif mv_method == 'ca' and mv_row and mv_col:
            result = run_ca(merged, mv_row, mv_col)
        elif mv_method == 'clustering' and len(mv_sel) >= 2:
            result = run_clustering(merged, mv_sel, mv_nk)
        else:
            flash("Paramètres insuffisants pour l'analyse.", 'warning')
            return redirect(url_for('multi_survey'))
    except Exception as e:
        flash(f"Erreur lors de l'analyse : {e}", 'danger')
        return redirect(url_for('multi_survey'))

    import io
    xlsx_bytes = _ms_build_xlsx(result, mv_method)
    method_labels = {'pca': 'ACP', 'mca': 'ACM', 'ca': 'AFC', 'clustering': 'Classification'}
    fname = f"analyse_multi_{method_labels.get(mv_method, mv_method)}.xlsx"
    return send_file(io.BytesIO(xlsx_bytes),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=fname)


# ─── MULTIVARIATE ─────────────────────────────────────────────────────────────

@app.route('/multivariate', methods=['GET', 'POST'])
def multivariate():
    df = get_dataframe()
    if df is None:
        return redirect(url_for('upload'))
    var_types = get_var_types(df)
    cat_vars = [v for v, t in var_types.items() if t == 'categorielle']
    num_vars = [v for v, t in var_types.items() if t == 'continue']
    # Include binary numeric for PCA/clustering
    bin_vars = [v for v, t in var_types.items() if t == 'binaire']
    num_vars_all = num_vars + bin_vars

    result = None
    method = None
    selected = []
    n_components = 2
    n_clusters = 3
    ca_row = ca_col = None

    if request.method == 'POST':
        method   = request.form.get('method')
        selected = request.form.getlist('variables')
        n_components = int(request.form.get('n_components', 2))
        n_clusters   = int(request.form.get('n_clusters', 3))
        ca_row = request.form.get('ca_row')
        ca_col = request.form.get('ca_col')
        # Persiste tous les paramètres pour le prochain GET
        session['multivariate_form'] = {
            'method': method, 'selected': selected,
            'n_components': n_components, 'n_clusters': n_clusters,
            'ca_row': ca_row, 'ca_col': ca_col,
        }
    else:
        saved = session.get('multivariate_form')
        if saved:
            method = saved.get('method')
            selected = saved.get('selected', [])
            n_components = saved.get('n_components', 2)
            n_clusters = saved.get('n_clusters', 3)
            ca_row = saved.get('ca_row')
            ca_col = saved.get('ca_col')
        else:
            # Première visite : ACP sur les variables numériques/binaires par
            # défaut — un premier coup d'œil structurel ne demande pas de
            # choix méthodologique préalable (contrairement à l'AFC, qui croise
            # 2 variables catégorielles précises et reste manuelle).
            # select_viable_variables() écarte les variables qui, combinées,
            # tomberaient à 0 ligne conjointement renseignée — cas fréquent
            # avec des variables mutuellement exclusives par logique de saut
            # (ex. « âge calculé » vs « âge saisi manuellement »), qui
            # feraient sinon échouer l'ACP silencieusement sur un simple GET.
            pca_vars = select_viable_variables(df, num_vars_all, min_rows=10) if len(num_vars_all) >= 2 else []
            if len(pca_vars) >= 2:
                method, selected = 'pca', pca_vars
            else:
                mca_vars = select_viable_variables(df, cat_vars, min_rows=10, numeric=False) if len(cat_vars) >= 2 else []
                if len(mca_vars) >= 2:
                    method, selected = 'mca', mca_vars

    # Filtre les variables qui n'existent plus dans le df actuel
    selected = [v for v in selected if v in df.columns]

    if method:
        try:
            if method == 'pca' and len(selected) >= 2:
                result = run_pca(df, selected, n_components)
            elif method == 'mca' and len(selected) >= 2:
                result = run_mca(df, selected, n_components)
            elif method == 'ca' and ca_row and ca_col and ca_row in df.columns and ca_col in df.columns:
                result = run_ca(df, ca_row, ca_col)
            elif method == 'clustering' and len(selected) >= 2:
                result = run_clustering(df, selected, n_clusters)
            elif request.method == 'POST':
                flash('Sélectionnez au moins 2 variables.', 'warning')
        except Exception as e:
            # Affiché même sur GET (pas seulement POST) : sinon l'analyse
            # automatique au premier chargement peut échouer sans qu'aucun
            # message n'explique pourquoi la page reste vide.
            flash(f'Erreur d\'analyse : {str(e)}', 'danger')

    return render_template('multivariate.html',
                           cat_vars=cat_vars,
                           num_vars=num_vars_all,
                           result=result, method=method,
                           selected_vars=selected,
                           n_components=n_components,
                           n_clusters=n_clusters,
                           ca_row=ca_row, ca_col=ca_col)


# ─── REPORT ───────────────────────────────────────────────────────────────────

@app.route('/report', methods=['GET', 'POST'])
def report():
    df = get_dataframe()
    if df is None:
        return redirect(url_for('upload'))
    var_types  = get_var_types(df)
    cat_vars   = [v for v, t in var_types.items() if t == 'categorielle']
    num_vars   = [v for v, t in var_types.items() if t == 'continue']
    bin_groups = detect_binary_groups(df)

    if request.method == 'POST':
        format_type        = request.form.get('report_format', 'pdf')
        title              = request.form.get('report_title', 'Rapport d\'analyse')
        author             = request.form.get('report_author', 'Analyste')
        selected_analyses  = request.form.getlist('analyses')
        if not selected_analyses:
            selected_analyses = ['descriptive']

        # Debug: log and show which analyses were selected (helps diagnose missing sections)
        app.logger.info('Selected analyses from form: %s', selected_analyses)
        flash(f"Analyses choisies : {', '.join(selected_analyses)}", 'info')
        selected_cat       = request.form.getlist('cat_vars')
        selected_num       = request.form.getlist('num_vars')
        selected_bin_grps  = request.form.getlist('bin_groups')
        crosstabs_row      = request.form.get('crosstabs_row')
        crosstabs_col      = request.form.get('crosstabs_col')

        extension = 'pdf' if format_type == 'pdf' else 'docx'
        report_path = os.path.join(
            app.config['REPORTS_FOLDER'],
            f'rapport_{uuid.uuid4().hex[:8]}.{extension}')
        try:
            user_comments = session.get('user_comments', {}) or {}
            # ── Multi-enquête : si demandé, on calcule les comparaisons à partir
            # des variables du panier (session) et on les passe au générateur.
            multi_survey_results = []
            multi_survey_meta = None
            if 'multi_survey' in selected_analyses:
                compared = list(session.get('ms_compared_vars', []) or [])
                surveys  = _load_multi_surveys()
                if compared and len(surveys) >= 2:
                    try:
                        mode = session.get('ms_mode', 'phakts')
                        merged = merge_surveys(surveys, mode=mode)
                        multi_survey_meta = {
                            'n_obs': int(len(merged)),
                            'n_vars': int(merged.shape[1] - 1),
                            'n_surveys': len(surveys),
                            'mode': mode,
                        }
                        for v in compared:
                            multi_survey_results.append(_compute_ms_comparison(merged, v))
                    except Exception as e:
                        flash(f"Erreur multi-enquête dans le rapport : {e}", 'warning')
                elif not compared:
                    flash("Aucune variable dans le panier multi-enquête — section ignorée.", 'warning')
                elif len(surveys) < 2:
                    flash("Moins de 2 enquêtes — section multi-enquête ignorée.", 'warning')

            if format_type == 'pdf':
                generate_pdf_report(
                    df, report_path,
                    title=title, author=author,
                    selected_analyses=selected_analyses,
                    cat_vars=selected_cat,
                    num_vars=selected_num,
                    bin_groups=[g for g in bin_groups if g['parent'] in selected_bin_grps],
                    crosstabs_row=crosstabs_row,
                    crosstabs_col=crosstabs_col,
                    user_comments=user_comments,
                    multi_survey_results=multi_survey_results,
                    multi_survey_meta=multi_survey_meta,
                )
            else:  # word
                generate_word_report(
                    df, report_path,
                    title=title, author=author,
                    selected_analyses=selected_analyses,
                    cat_vars=selected_cat,
                    num_vars=selected_num,
                    bin_groups=[g for g in bin_groups if g['parent'] in selected_bin_grps],
                    crosstabs_row=crosstabs_row,
                    crosstabs_col=crosstabs_col,
                    user_comments=user_comments,
                    multi_survey_results=multi_survey_results,
                    multi_survey_meta=multi_survey_meta,
                )
            return send_file(report_path, as_attachment=True,
                             download_name=f'{title.replace(" ","_")}.{extension}')
        except Exception as e:
            flash(f'Erreur génération rapport : {str(e)}', 'danger')

    # Pré-remplir le formulaire avec les sélections faites dans les onglets
    # descriptive / crosstabs (sinon : tout cocher par défaut, comportement précédent).
    desc_state = session.get('descriptive_form', {})
    ct_state   = session.get('crosstabs_form', {})
    default_cat = desc_state.get('cat') if desc_state else cat_vars
    default_num = desc_state.get('num') if desc_state else num_vars
    default_bin = desc_state.get('bin') if desc_state else [g['parent'] for g in bin_groups]
    default_ct_row = ct_state.get('row_var', '') if ct_state else ''
    default_ct_col = ct_state.get('col_var', '') if ct_state else ''

    return render_template('report.html',
                           cat_vars=cat_vars, num_vars=num_vars,
                           bin_groups=bin_groups,
                           default_cat=default_cat,
                           default_num=default_num,
                           default_bin=default_bin,
                           default_ct_row=default_ct_row,
                           default_ct_col=default_ct_col,
                           meta=session.get('data_meta', {}))



# ─── KOBO CONNECT ─────────────────────────────────────────────────────────────

@app.route('/kobo/connect', methods=['GET', 'POST'])
def kobo_connect():
    """Page de connexion KoboToolbox — saisie et validation du token."""
    if request.method == 'POST':
        token           = request.form.get('token', '').strip()
        custom_instance = request.form.get('custom_instance', '').strip()
        if not token:
            flash('Veuillez saisir votre token API.', 'warning')
            return redirect(url_for('kobo_connect'))
        result = validate_token(token, custom_instance=custom_instance or None)
        if result['valid']:
            session['kobo_token']    = token
            session['kobo_username'] = result.get('username', '')
            session['kobo_instance'] = result.get('instance', '')
            kobo_track.resume(token, result.get('instance'))
            flash(
                f"Connecté en tant que <strong>{result.get('username','—')}</strong> "
                f"— Instance : <strong>{result.get('instance_label','KoboToolbox')}</strong>.",
                'success'
            )
            return redirect(url_for('kobo_assets'))
        else:
            flash(f"Connexion échouée : {result['error']}", 'danger')
    already = bool(session.get('kobo_token'))
    return render_template('kobo_connect.html', already=already,
                           username=session.get('kobo_username', ''),
                           instance=session.get('kobo_instance', ''))


@app.route('/kobo/assets')
def kobo_assets():
    """Liste les formulaires disponibles pour le token en session."""
    token    = session.get('kobo_token')
    instance = session.get('kobo_instance')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    result = list_assets(token, instance=instance)
    if not result['success']:
        flash(f"Impossible de lister les formulaires : {result['error']}", 'danger')
        return redirect(url_for('kobo_connect'))
    # Mémoriser l'instance si elle a été détectée
    if result.get('instance'):
        session['kobo_instance'] = result['instance']
    return render_template('kobo_connect.html',
                           assets=result['assets'],
                           total=result['total'],
                           username=session.get('kobo_username', ''),
                           instance=session.get('kobo_instance', ''),
                           already=True)


@app.route('/kobo/load', methods=['POST'])
def kobo_load():
    """
    Charge les soumissions du formulaire sélectionné depuis KoboToolbox.
    Sauvegarde le DataFrame en Excel et redirige vers data_preview
    — même mécanique que le flux upload fichier.
    """
    token = session.get('kobo_token')
    uid   = request.form.get('uid', '').strip()
    name  = request.form.get('name', 'Formulaire KoboToolbox')
    if not token:
        flash('Session expirée. Reconnectez-vous à KoboToolbox.', 'warning')
        return redirect(url_for('kobo_connect'))
    if not uid:
        flash('UID de formulaire manquant.', 'danger')
        return redirect(url_for('kobo_assets'))

    kobo_sync.stop()  # un nouveau formulaire est chargé : le polling précédent ne s'applique plus
    result = kobo_load_data(token, uid)
    if not result['success']:
        flash(f"Erreur de chargement : {result['error']}", 'danger')
        return redirect(url_for('kobo_assets'))

    df        = result['df']
    fname     = f"kobo_{uuid.uuid4().hex[:8]}.xlsx"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    df.to_excel(save_path, index=False, engine='openpyxl')

    session['data_path']           = save_path
    session['data_meta']           = summarize_dataframe(df)
    session['data_meta']['source'] = 'KoboToolbox'
    session['data_meta']['name']   = name
    session['kobo_uid']            = uid
    session['kobo_asset_name']     = name
    session['original_filename']   = f"{name}.xlsx"
    session.pop('child_path', None)

    flash(
        f"<strong>{name}</strong> chargé avec succès — "
        f"{result['n_obs']} soumissions × {result['n_vars']} variables analysables.",
        'success'
    )
    return redirect(url_for('data_preview'))


@app.route('/kobo/refresh')
def kobo_refresh():
    """
    Re-télécharge les dernières soumissions du formulaire en session.
    Accessible depuis data_preview via le bouton Rafraîchir.
    """
    token = session.get('kobo_token')
    uid   = session.get('kobo_uid')
    name  = session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    if not token or not uid:
        flash('Aucun formulaire KoboToolbox actif en session.', 'warning')
        return redirect(url_for('kobo_connect'))

    result = kobo_load_data(token, uid)
    if not result['success']:
        flash(f"Erreur de rafraîchissement : {result['error']}", 'danger')
        return redirect(url_for('data_preview'))

    df = result['df']
    old_path = session.get('data_path')
    if old_path and os.path.exists(old_path):
        try: os.remove(old_path)
        except Exception: pass

    fname     = f"kobo_{uuid.uuid4().hex[:8]}.xlsx"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    df.to_excel(save_path, index=False, engine='openpyxl')

    session['data_path']           = save_path
    session['data_meta']           = summarize_dataframe(df)
    session['data_meta']['source'] = 'KoboToolbox'
    session['data_meta']['name']   = name
    kobo_sync.set_baseline(result['n_obs'])

    flash(
        f"Données rafraîchies — {result['n_obs']} soumissions "
        f"× {result['n_vars']} variables.",
        'success'
    )
    return redirect(url_for('data_preview'))


@app.route('/kobo/disconnect')
def kobo_disconnect():
    """Efface le token KoboToolbox de la session."""
    kobo_sync.stop()
    kobo_track.clear_all()
    for k in ('kobo_token', 'kobo_username', 'kobo_uid', 'kobo_asset_name', 'kobo_instance'):
        session.pop(k, None)
    flash('Déconnecté de KoboToolbox.', 'info')
    return redirect(url_for('kobo_connect'))


# ─── Suivi multi-formulaires (plusieurs enquêtes en parallèle) ────────────────
#
# Complète /kobo/sync/* (qui suit UN formulaire, celui en cours d'analyse) par
# un suivi simultané de PLUSIEURS formulaires — sondage léger (compteur de
# soumissions uniquement, jamais les données complètes) pour rester rapide
# même avec plusieurs formulaires suivis en parallèle.

@app.route('/suivi')
def suivi():
    """Page de suivi multi-formulaires."""
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox pour suivre des formulaires.", 'warning')
        return redirect(url_for('kobo_connect'))
    instance = session.get('kobo_instance')
    assets_result = list_assets(token, instance=instance)
    assets = assets_result.get('assets', []) if assets_result.get('success') else []
    tracked = kobo_track.list_tracked()
    tracked_uids = {t['uid'] for t in tracked}
    guessed_types = {a['uid']: ref_data.guess_form_type(a['name']) for a in assets}

    # Même logique d'alerte que « Complétude nationale » (validate_mapping) :
    # un type manuellement choisi qui ne correspond pas au nom du formulaire,
    # ou le même type SPAD assigné à 2 formulaires suivis différents, sont
    # des erreurs de sélection qui passeraient sinon inaperçues.
    suivi_avertissements = []
    types_vus = {}
    for t in tracked:
        if not t.get('form_type'):
            continue
        types_vus.setdefault(t['form_type'], []).append(t['name'])
        guess = ref_data.guess_form_type(t['name'])
        if guess and guess != t['form_type']:
            suivi_avertissements.append(
                f"« {t['name']} » est suivi comme {t['form_type']}, mais son nom correspond "
                f"plutôt à {guess} — vérifiez la sélection."
            )
    for code, noms in types_vus.items():
        if len(noms) > 1:
            suivi_avertissements.append(
                f"{code} est assigné à {len(noms)} formulaires suivis différents "
                f"({', '.join(noms)}) — un seul devrait normalement porter ce type."
            )

    return render_template(
        'suivi.html',
        assets=assets,
        tracked_uids=tracked_uids,
        assets_error=None if assets_result.get('success') else assets_result.get('error'),
        form_codes=ref_data.FORM_CODES,
        form_labels=ref_data.FORM_LABELS,
        guessed_types=guessed_types,
        suivi_avertissements=suivi_avertissements,
        ai_available=ai_form_assist.available(),
    )


def _parse_target_form(request_form):
    """Lit uid/name/target/form_type communs à /suivi/add et /suivi/target."""
    target_raw = (request_form.get('target') or '').strip()
    form_type = (request_form.get('form_type') or '').strip() or None
    if form_type and form_type not in ref_data.FORM_CODES:
        form_type = None
    target = None
    if target_raw:
        try:
            target = int(target_raw)
            if target <= 0:
                target = None
        except ValueError:
            return None, None, "Cible invalide (nombre entier attendu)."
    return target, form_type, None


def _sync_completude_mapping(form_type, uid):
    """Répercute automatiquement dans la correspondance de Complétude
    nationale (form_mapping.py) le type assigné à un formulaire suivi —
    qu'il vienne de la détection par préfixe, du bouton IA ou d'un choix
    manuel. Évite l'étape séparée « aller réassocier à la main dans
    Complétude » à chaque fois qu'un formulaire est ajouté ou retypé dans
    Suivi ; sans effet si form_type est vide (type « Libre »)."""
    if not form_type:
        return
    mapping = form_mapping.load()
    if mapping.get(form_type) != uid:
        mapping[form_type] = uid
        form_mapping.save(mapping)


@app.route('/suivi/add', methods=['POST'])
def suivi_add():
    token = session.get('kobo_token')
    instance = session.get('kobo_instance')
    if not token:
        return jsonify({"success": False, "error": "Non connecté à KoboToolbox."}), 400
    uid = (request.form.get('uid') or '').strip()
    name = (request.form.get('name') or 'Formulaire').strip()
    if not uid:
        return jsonify({"success": False, "error": "Formulaire manquant."}), 400
    target, form_type, err = _parse_target_form(request.form)
    if err:
        return jsonify({"success": False, "error": err}), 400
    kobo_track.add(token, instance, uid, name, target=target, form_type=form_type)
    _sync_completude_mapping(form_type, uid)
    return jsonify({"success": True, "tracked": kobo_track.list_tracked()})


@app.route('/suivi/remove', methods=['POST'])
def suivi_remove():
    uid = (request.form.get('uid') or '').strip()
    kobo_track.remove(uid)
    return jsonify({"success": True, "tracked": kobo_track.list_tracked()})


@app.route('/suivi/target', methods=['POST'])
def suivi_target():
    uid = (request.form.get('uid') or '').strip()
    target, form_type, err = _parse_target_form(request.form)
    if err:
        return jsonify({"success": False, "error": err}), 400
    if not kobo_track.is_tracked(uid):
        return jsonify({"success": False, "error": "Formulaire non suivi."}), 400
    kobo_track.set_target(uid, target=target, form_type=form_type)
    _sync_completude_mapping(form_type, uid)
    return jsonify({"success": True, "tracked": kobo_track.list_tracked()})


@app.route('/suivi/ai_suggest', methods=['POST'])
def suivi_ai_suggest():
    """Propose un type de formulaire (registre existant ou nouveau) et une
    cible via l'IA (modules/ai_form_assist.py) — pré-remplit uniquement les
    champs du formulaire de suivi côté client, n'enregistre jamais rien lui-
    même. Renvoie {'available': False} sans erreur si ANTHROPIC_API_KEY
    n'est pas configurée (le bouton reste alors masqué côté template)."""
    name = (request.form.get('name') or '').strip()
    if not name:
        return jsonify({'available': False, 'error': 'Nom de formulaire manquant.'}), 400
    return jsonify(ai_form_assist.suggest_for_kobo_form(name))


def _enrich_tracked_district_reel(tracked):
    """Ajoute la cible réelle par district (voir cp.district_reel) aux
    formulaires suivis reconnus comme un code SPAD officiel (form_type) —
    réutilise le dernier calcul de « Complétude nationale » en cache s'il
    couvre ce formulaire. Reste None pour les formulaires Kobo non-SPAD :
    kobo_track.py ne connaît volontairement aucun référentiel régions/
    districts pour eux (sondage léger, voir modules/kobo_track.py)."""
    cached = _load_completude_cache()
    district_table = cached.get('district') if cached else None
    for t in tracked:
        t['district_reel'] = None
        if district_table and t.get('form_type'):
            t['district_reel'] = cp.district_reel(district_table, t['form_type'])
    return tracked


@app.route('/suivi/status')
def suivi_status():
    return jsonify({"tracked": _enrich_tracked_district_reel(kobo_track.list_tracked())})


@app.route('/suivi/analyser/<uid>')
def suivi_analyser(uid):
    """Charge ce formulaire suivi dans le flux d'analyse mono-enquête existant."""
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    tracked = {t['uid']: t for t in kobo_track.list_tracked()}
    name = tracked.get(uid, {}).get('name', 'Formulaire KoboToolbox')
    kobo_sync.stop()
    result = kobo_load_data(token, uid)
    if not result['success']:
        flash(f"Erreur de chargement : {result['error']}", 'danger')
        return redirect(url_for('suivi'))
    df = result['df']
    fname = f"kobo_{uuid.uuid4().hex[:8]}.xlsx"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    df.to_excel(save_path, index=False, engine='openpyxl')
    session['data_path']           = save_path
    session['data_meta']           = summarize_dataframe(df)
    session['data_meta']['source'] = 'KoboToolbox'
    session['data_meta']['name']   = name
    session['kobo_uid']            = uid
    session['kobo_asset_name']     = name
    session.pop('child_path', None)
    flash(f"<strong>{name}</strong> chargé pour analyse — {result['n_obs']} soumissions × {result['n_vars']} variables.", 'success')
    return redirect(url_for('data_preview'))


# ─── Complétude nationale (référentiel + moteur de complétude) ────────────────
#
# Reproduit le tableau de bord de complétude façon spadapp-zeta.vercel.app :
# l'utilisateur associe chacun des 7 formulaires SPAD à son formulaire Kobo
# réel (un mapping, pas une détection automatique — les UID Kobo ne portent
# aucune information sur le formulaire), puis « Calculer » tire les données
# de chaque formulaire mappé et calcule reçu/cible/taux/statut via
# modules/completeness.py, au regard du référentiel modules/reference_data.py.

def _load_completude_cache():
    """Relit le résultat de calcul depuis le fichier JSON référencé en session.

    Les tables région/district (12 × 7 formulaires × plusieurs champs)
    dépassent largement la limite d'un cookie de session Flask (~4 Ko) — donc,
    comme pour le jeu de données analysé (session['data_path']), seul le
    CHEMIN du fichier de résultat est stocké en session, jamais son contenu.

    Filet de sécurité : si le cookie de session ne référence pas (ou plus)
    de fichier valide — observé en usage réel dans le contexte iframe de
    l'app desktop Electron, cause exacte non confirmée — on retombe sur le
    fichier completude_*.json le plus récent du dossier d'upload plutôt que
    de perdre un calcul qui a pourtant réussi côté serveur. Une incohérence
    multi-utilisateur n'est pas un risque ici (app mono-utilisateur locale).
    """
    path = session.get('completude_path')
    if path and os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass

    try:
        candidates = [
            os.path.join(app.config['UPLOAD_FOLDER'], f)
            for f in os.listdir(app.config['UPLOAD_FOLDER'])
            if f.startswith('completude_') and f.endswith('.json')
        ]
        if not candidates:
            return None
        latest = max(candidates, key=os.path.getmtime)
        with open(latest, 'r', encoding='utf-8') as f:
            data = json.load(f)
        session['completude_path'] = latest  # resynchronise la session pour les prochaines requêtes
        return data
    except Exception:
        return None


def _kobo_credentials():
    """Jeton/instance Kobo à utiliser pour la Complétude nationale : la session
    de l'utilisateur connecté a priorité, sinon le jeton serveur
    (KOBO_API_TOKEN, KOBO_INSTANCE) — pour que ce tableau reste à jour sans
    qu'aucun rôle (Data ou Invité) n'ait à se connecter à KoboToolbox."""
    token = session.get('kobo_token') or (os.environ.get('KOBO_API_TOKEN') or '').strip() or None
    instance = session.get('kobo_instance') or (os.environ.get('KOBO_INSTANCE') or '').strip() or None
    return token, instance


@app.route('/completude')
def completude():
    # Le menu de correspondance liste les formulaires déjà ajoutés dans
    # « Suivi multi-formulaires » (kobo_track.list_tracked()) plutôt qu'un
    # appel indépendant à l'API Kobo — la sélection reste cohérente entre
    # les deux écrans, et accessible sans session Kobo (y compris au rôle
    # 'invite', qui n'en a jamais). Ajoutez d'abord un formulaire dans
    # Suivi pour qu'il devienne sélectionnable ici.
    assets = [{'uid': t['uid'], 'name': t['name'], 'submission_count': t.get('count')}
              for t in kobo_track.list_tracked()]
    assets_error = None
    mapping = form_mapping.load()
    cached = _load_completude_cache()
    mapping_erreurs, mapping_avertissements = ref_data.validate_mapping(mapping, assets)
    district_reel = None
    if cached and cached.get('district'):
        district_reel = {code: cp.district_reel(cached['district'], code) for code in ref_data.FORM_CODES}
    return render_template(
        'completude.html',
        assets=assets,
        assets_error=assets_error,
        mapping=mapping,
        form_codes=ref_data.FORM_CODES,
        form_labels=ref_data.FORM_LABELS,
        result=cached['national'] if cached else None,
        district_reel=district_reel,
        computed_at=session.get('completude_computed_at'),
        mapping_erreurs=mapping_erreurs,
        mapping_avertissements=mapping_avertissements,
    )


@app.route('/completude/regions')
def completude_regions():
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    return render_template(
        'completude_table.html',
        title='Complétude par région', echelon='région',
        rows=cached['region'], form_codes=ref_data.FORM_CODES, form_labels=ref_data.FORM_LABELS,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/districts')
def completude_districts():
    # Lecture seule d'un calcul déjà en cache — aucun besoin d'une session
    # Kobo active (accessible au rôle 'invite', qui n'en a jamais).
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    return render_template(
        'completude_table.html',
        title='Complétude par district', echelon='district',
        rows=cached['district'], form_codes=ref_data.FORM_CODES, form_labels=ref_data.FORM_LABELS,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/enqueteurs')
def completude_enqueteurs():
    # Lecture seule d'un calcul déjà en cache — accessible au rôle 'invite'.
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    return render_template(
        'completude_table.html',
        title='Complétude par enquêteur', echelon='enquêteur', sous_titre_label='District',
        rows=cached['enqueteur'], form_codes=list(cp.enqueteur_forms()), form_labels=ref_data.FORM_LABELS,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/superviseurs')
def completude_superviseurs():
    # Lecture seule d'un calcul déjà en cache — accessible au rôle 'invite'.
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    return render_template(
        'completude_table.html',
        title='Complétude par superviseur', echelon='superviseur', sous_titre_label='District',
        rows=cached['superviseur'], form_codes=list(cp.superviseur_forms()), form_labels=ref_data.FORM_LABELS,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/enqueteurs/<enq_code>')
def completude_enqueteur_detail(enq_code):
    # Lecture seule d'un calcul déjà en cache — accessible au rôle 'invite'.
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    enq = cached['enqueteur'].get(enq_code)
    if not enq:
        flash('Enquêteur introuvable.', 'warning')
        return redirect(url_for('completude_enqueteurs'))

    etablissements = {c: e for c, e in cached['etablissement'].items() if e['enqueteur_code'] == enq_code}
    etab_codes = set(etablissements)
    zeros = [a for a in cached.get('anomalies_zero', []) if a.get('etablissement_code') in etab_codes]
    excess = [a for a in cached.get('anomalies_excess', []) if a.get('etablissement_code') in etab_codes]

    return render_template(
        'completude_enqueteur_detail.html',
        enq_code=enq_code, enq=enq, etablissements=etablissements,
        form_codes=list(cp.enqueteur_forms()), form_labels=ref_data.FORM_LABELS,
        anomalies_zero=zeros, anomalies_excess=excess,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/superviseurs/<sup_code>')
def completude_superviseur_detail(sup_code):
    # Lecture seule d'un calcul déjà en cache — accessible au rôle 'invite'.
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    sup = cached['superviseur'].get(sup_code)
    if not sup:
        flash('Superviseur introuvable.', 'warning')
        return redirect(url_for('completude_superviseurs'))

    district_code = sup['sous_titre']
    district = cached['district'].get(district_code)
    etablissements = {c: e for c, e in cached['etablissement'].items() if e['district_code'] == district_code}
    zeros, excess = _anomalies_scope(cached, district_code=district_code)

    superviseur_forms = cp.superviseur_forms()
    return render_template(
        'completude_superviseur_detail.html',
        sup_code=sup_code, sup=sup, district_code=district_code, district=district,
        etablissements=etablissements,
        superviseur_form_codes=list(superviseur_forms),
        # Parmi les formulaires du volet RDM, ceux au grain établissement
        # (motif F02) — la table « détail établissement » ci-dessous s'y
        # adapte plutôt que de supposer qu'il n'y en a toujours qu'un (F02).
        superviseur_etab_form_codes=[c for c in superviseur_forms if c in cp.etablissement_forms()],
        form_labels=ref_data.FORM_LABELS,
        anomalies_zero=zeros, anomalies_excess=excess,
        computed_at=session.get('completude_computed_at'),
    )


def _anomalies_scope(cached, **scope):
    """Filtre les anomalies (0 % / excédent) déjà calculées selon un ou
    plusieurs codes (region_code / district_code / etablissement_code)."""
    def _match(a):
        return all(a.get(k) == v for k, v in scope.items())
    zeros = [a for a in cached.get('anomalies_zero', []) if _match(a)]
    excess = [a for a in cached.get('anomalies_excess', []) if _match(a)]
    return zeros, excess


@app.route('/completude/regions/<region_code>')
def completude_region_detail(region_code):
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    region = cached['region'].get(region_code)
    if not region:
        flash('Région introuvable.', 'warning')
        return redirect(url_for('completude_regions'))

    districts = {c: d for c, d in cached['district'].items() if d['region_code'] == region_code}
    zeros, excess = _anomalies_scope(cached, region_code=region_code)

    return render_template(
        'completude_region_detail.html',
        region_code=region_code, region=region, districts=districts,
        form_codes=ref_data.FORM_CODES, form_labels=ref_data.FORM_LABELS,
        anomalies_zero=zeros, anomalies_excess=excess,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/districts/<district_code>')
def completude_district_detail(district_code):
    # Lecture seule d'un calcul déjà en cache — accessible au rôle 'invite'.
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    district = cached['district'].get(district_code)
    if not district:
        flash('District introuvable.', 'warning')
        return redirect(url_for('completude_districts'))

    etablissements = {c: e for c, e in cached['etablissement'].items() if e['district_code'] == district_code}
    enqueteurs = {c: e for c, e in cached['enqueteur'].items() if e['sous_titre'] == district_code}
    superviseur = next(((c, s) for c, s in cached['superviseur'].items() if s['sous_titre'] == district_code), None)
    zeros, excess = _anomalies_scope(cached, district_code=district_code)

    return render_template(
        'completude_district_detail.html',
        district_code=district_code, district=district, etablissements=etablissements,
        enqueteurs=enqueteurs, superviseur=superviseur,
        form_codes=ref_data.FORM_CODES,
        etablissement_form_codes=list(cp.etablissement_forms()),
        enqueteur_form_codes=list(cp.enqueteur_forms()), superviseur_form_codes=list(cp.superviseur_forms()),
        form_labels=ref_data.FORM_LABELS,
        anomalies_zero=zeros, anomalies_excess=excess,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/etablissements/<etablissement_code>')
def completude_etablissement_detail(etablissement_code):
    # Lecture seule d'un calcul déjà en cache — accessible au rôle 'invite'.
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    etab = cached['etablissement'].get(etablissement_code)
    if not etab:
        flash('Établissement introuvable.', 'warning')
        return redirect(url_for('completude_districts'))

    enqueteur = cached['enqueteur'].get(etab['enqueteur_code'])
    superviseur = next(((c, s) for c, s in cached['superviseur'].items() if s['sous_titre'] == etab['district_code']), None)
    zeros, excess = _anomalies_scope(cached, etablissement_code=etablissement_code)

    return render_template(
        'completude_etablissement_detail.html',
        etablissement_code=etablissement_code, etab=etab,
        enqueteur_code=etab['enqueteur_code'], enqueteur=enqueteur, superviseur=superviseur,
        form_codes=list(cp.etablissement_forms()), form_labels=ref_data.FORM_LABELS,
        anomalies_zero=zeros, anomalies_excess=excess,
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/anomalies')
def completude_anomalies():
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    return render_template(
        'completude_anomalies.html',
        zeros=cached.get('anomalies_zero', []),
        excess=cached.get('anomalies_excess', []),
        computed_at=session.get('completude_computed_at'),
    )


@app.route('/completude/graphiques')
def completude_graphiques():
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))

    statut_counts = {'zero': 0, 'en_cours': 0, 'cible': 0, 'verifier': 0}
    for r in cached.get('export', []):
        if r['statut'] in statut_counts:
            statut_counts[r['statut']] += 1

    return render_template(
        'completude_graphiques.html',
        district_table=cached.get('district', {}),
        form_codes=ref_data.FORM_CODES,
        form_labels=ref_data.FORM_LABELS,
        statut_counts=statut_counts,
        historique=tendance.load_history(30),
        computed_at=session.get('completude_computed_at'),
    )


def _export_filename(ext):
    return f"spad_completude_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.{ext}"


@app.route('/completude/export.csv')
def completude_export_csv():
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    rows = cached.get('export', [])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['région', 'district', 'établissement', 'formulaire', 'libellé formulaire',
                      'cible', 'reçu', 'taux (%)', 'statut'])
    for r in rows:
        writer.writerow([r['region'], r['district'], r['unite'], r['formulaire'], r['formulaire_label'],
                          r['cible'], r['recu'], r['taux'] if r['taux'] is not None else '', r['statut']])

    resp = Response(buf.getvalue(), mimetype='text/csv; charset=utf-8')
    resp.headers['Content-Disposition'] = f'attachment; filename="{_export_filename("csv")}"'
    return resp


@app.route('/completude/export.xlsx')
def completude_export_xlsx():
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    rows = cached.get('export', [])

    df = pd.DataFrame(rows, columns=['region', 'district', 'unite', 'formulaire', 'formulaire_label',
                                      'cible', 'recu', 'taux', 'statut'])
    df.columns = ['Région', 'District', 'Établissement', 'Formulaire', 'Libellé formulaire',
                  'Cible', 'Reçu', 'Taux (%)', 'Statut']
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Complétude')
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=_export_filename('xlsx'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/completude/export.docx')
def completude_export_docx():
    """Rapport de complétude au format Word — voir modules/completude_report.py
    (structure inspirée d'un compte rendu de débriefing terrain fourni par
    l'utilisateur : indicateurs clés, points d'alerte, complétude par
    formulaire/district, actions suggérées à partir des anomalies détectées)."""
    cached = _load_completude_cache()
    if not cached:
        flash("Calculez d'abord la complétude depuis la page « Complétude nationale ».", 'warning')
        return redirect(url_for('completude'))
    ref = ref_data.load()
    docx_bytes = completude_report.build_docx(
        cached, ref, computed_at=session.get('completude_computed_at'))
    buf = io.BytesIO(docx_bytes)
    return send_file(
        buf, as_attachment=True, download_name=_export_filename('docx'),
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )


# ─── Projets d'enquête génériques (Kobo/ODK — hors référentiel SPAD) ──────────
#
# Généralise le suivi de complétude (reçu/cible/taux/statut) à n'importe
# quelle autre enquête, à partir d'un fichier de référence fourni par
# l'utilisateur (colonnes minimales : code, cible) plutôt que du référentiel
# SPAD figé. Voir modules/projets.py pour le détail.

@app.route('/projets')
def projets_liste():
    token = session.get('kobo_token')
    instance = session.get('kobo_instance')
    assets = []
    assets_error = None
    if token:
        assets_result = list_assets(token, instance=instance)
        assets = assets_result.get('assets', []) if assets_result.get('success') else []
        assets_error = None if assets_result.get('success') else assets_result.get('error')
    return render_template(
        'projets.html',
        projets=proj.list_projets(),
        kobo_connected=bool(token),
        assets=assets,
        assets_error=assets_error,
    )


@app.route('/projets/creer', methods=['POST'])
def projets_creer():
    nom = (request.form.get('nom') or '').strip()
    champ_unite = (request.form.get('champ_unite') or '').strip()
    kobo_uid = (request.form.get('kobo_uid') or '').strip() or None
    kobo_name = (request.form.get('kobo_name') or '').strip() or None

    if not nom or not champ_unite:
        flash("Nom du projet et champ d'identification de l'unité obligatoires.", 'warning')
        return redirect(url_for('projets_liste'))
    if 'reference' not in request.files or request.files['reference'].filename == '':
        flash('Fichier de référence manquant (colonnes minimales : code, cible).', 'warning')
        return redirect(url_for('projets_liste'))

    f = request.files['reference']
    if not allowed_file(f.filename):
        flash('Format non supporté pour le fichier de référence (.xlsx / .xls attendu).', 'danger')
        return redirect(url_for('projets_liste'))

    try:
        entry = proj.create_projet(
            nom, f, champ_unite,
            kobo_uid=kobo_uid, kobo_name=kobo_name,
            kobo_instance=session.get('kobo_instance'),
        )
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('projets_liste'))

    flash(f"Projet « {entry['nom']} » créé — {entry['n_unites']} unité(s) dans le référentiel.", 'success')
    return redirect(url_for('projets_detail', projet_id=entry['id']))


@app.route('/projets/<projet_id>')
def projets_detail(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    result = proj.load_result(projet_id)
    token = session.get('kobo_token')
    instance = session.get('kobo_instance')
    assets = []
    if token:
        assets_result = list_assets(token, instance=instance)
        assets = assets_result.get('assets', []) if assets_result.get('success') else []
    return render_template(
        'projet_detail.html',
        p=p, assets=assets, kobo_connected=bool(token),
        unites=result['unites'] if result else None,
        groupes=result['groupes'] if result else None,
        analyse_resume=proj.load_analyse_resume(projet_id),
        rapport_disponible=bool(proj.rapport_path(projet_id)),
    )


@app.route('/projets/<projet_id>/lier', methods=['POST'])
def projets_lier(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    kobo_uid = (request.form.get('kobo_uid') or '').strip()
    kobo_name = (request.form.get('kobo_name') or '').strip()
    if kobo_uid:
        proj.update_kobo_link(projet_id, kobo_uid, kobo_name, kobo_instance=session.get('kobo_instance'))
        flash(f"Formulaire KoboToolbox « {kobo_name} » associé au projet.", 'success')
    return redirect(url_for('projets_detail', projet_id=projet_id))


@app.route('/projets/<projet_id>/calculer', methods=['POST'])
def projets_calculer(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    if not p.get('kobo_uid'):
        flash("Associez d'abord un formulaire KoboToolbox à ce projet.", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    res = kobo_load_data(token, p['kobo_uid'], instance=p.get('kobo_instance') or session.get('kobo_instance'))
    if not res.get('success'):
        flash(f"Erreur de chargement : {res.get('error')}", 'danger')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    reference = proj.load_reference(projet_id)
    unites = proj.unit_completeness(reference, p['champ_unite'], res['df'])
    groupes = proj.group_completeness(unites)
    proj.save_result(projet_id, unites, groupes)

    flash(f"Complétude recalculée — {res['n_obs']} soumissions traitées.", 'success')
    return redirect(url_for('projets_detail', projet_id=projet_id))


# ─── Analyse épidémiologique générique (XLSForm → dictionnaire → rapport) ─────
#
# Étend un projet avec un XLSForm optionnel : dictionnaire de variables
# auto-généré (suggestions éditables), puis statistiques univariées, scores
# composites, stratification et rapport Word — pour n'importe quel
# questionnaire, pas seulement les formulaires SPAD. Voir modules/
# xlsform_dictionary.py et modules/enquete_analyse.py.

@app.route('/projets/<projet_id>/xlsform', methods=['POST'])
def projets_xlsform(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    if 'xlsform' not in request.files or request.files['xlsform'].filename == '':
        flash("Fichier XLSForm manquant (feuilles 'survey' et 'choices' attendues).", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))
    f = request.files['xlsform']
    if not allowed_file(f.filename):
        flash('Format non supporté (.xlsx / .xls attendu).', 'danger')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    try:
        n = proj.attach_xlsform(projet_id, f)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    flash(f"XLSForm chargé — dictionnaire généré ({n} variables). Révisez-le avant de lancer l'analyse.",
          'success')
    return redirect(url_for('projets_dictionnaire', projet_id=projet_id))


@app.route('/projets/<projet_id>/dictionnaire')
def projets_dictionnaire(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    dic = proj.load_dictionnaire(projet_id)
    if dic is None:
        flash("Aucun XLSForm attaché à ce projet.", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))
    return render_template(
        'projet_dictionnaire.html', p=p,
        lignes=dic.to_dict('records'),
        domaines_existants=sorted({d for d in dic['domaine'] if d}),
    )


@app.route('/projets/<projet_id>/dictionnaire/save', methods=['POST'])
def projets_dictionnaire_save(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    dic = proj.load_dictionnaire(projet_id)
    if dic is None:
        flash("Aucun XLSForm attaché à ce projet.", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    for i in dic.index:
        nom = dic.at[i, 'nom']
        dic.at[i, 'role'] = request.form.get(f'role_{nom}', dic.at[i, 'role'])
        dic.at[i, 'domaine'] = (request.form.get(f'domaine_{nom}') or '').strip()
        dic.at[i, 'inclure_score_composite'] = request.form.get(f'inclure_{nom}') == 'on'
        dic.at[i, 'sens_item'] = request.form.get(f'sens_{nom}', '')
        dic.at[i, 'valeurs_favorables'] = (request.form.get(f'favorables_{nom}') or '').strip()

    proj.save_dictionnaire(projet_id, dic)
    flash('Dictionnaire enregistré.', 'success')
    return redirect(url_for('projets_dictionnaire', projet_id=projet_id))


@app.route('/projets/<projet_id>/analyser', methods=['POST'])
def projets_analyser(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    if proj.load_dictionnaire(projet_id) is None:
        flash("Attachez et validez un dictionnaire avant de lancer l'analyse.", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))
    token = session.get('kobo_token')
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('kobo_connect'))
    if not p.get('kobo_uid'):
        flash("Associez d'abord un formulaire KoboToolbox à ce projet.", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    res = kobo_load_data(token, p['kobo_uid'], instance=p.get('kobo_instance') or session.get('kobo_instance'))
    if not res.get('success'):
        flash(f"Erreur de chargement : {res.get('error')}", 'danger')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    try:
        proj.run_analysis(projet_id, res['df'])
    except Exception as e:
        flash(f"Erreur pendant l'analyse : {e}", 'danger')
        return redirect(url_for('projets_detail', projet_id=projet_id))

    flash(f"Analyse terminée — {res['n_obs']} soumissions traitées. Rapport disponible au téléchargement.",
          'success')
    return redirect(url_for('projets_detail', projet_id=projet_id))


@app.route('/projets/<projet_id>/rapport.docx')
def projets_rapport_docx(projet_id):
    p = proj.get_projet(projet_id)
    if not p:
        flash('Projet introuvable.', 'warning')
        return redirect(url_for('projets_liste'))
    path = proj.rapport_path(projet_id)
    if not path:
        flash("Aucun rapport disponible — lancez d'abord l'analyse.", 'warning')
        return redirect(url_for('projets_detail', projet_id=projet_id))
    fname = f"rapport_analyse_{p['nom'].replace(' ', '_')}.docx"
    return send_file(
        path, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )


@app.route('/projets/<projet_id>/supprimer', methods=['POST'])
def projets_supprimer(projet_id):
    proj.delete_projet(projet_id)
    flash('Projet supprimé.', 'info')
    return redirect(url_for('projets_liste'))


@app.route('/completude/mapper', methods=['POST'])
def completude_mapper():
    # On part du mapping existant (et non d'un dict vide) : cette route ne
    # doit modifier que les codes réellement présents dans le formulaire
    # soumis (les formulaires actifs au moment de l'affichage de la page).
    # Un code temporairement désactivé dans le registre (donc absent de
    # ref_data.FORM_CODES à cet instant) ne doit jamais perdre son
    # association déjà enregistrée simplement parce que ce bouton a été
    # cliqué — sinon désactiver puis réactiver un formulaire effacerait sa
    # correspondance Kobo sans qu'aucun utilisateur ne l'ait demandé.
    mapping = form_mapping.load()
    n_soumis = 0
    for code in ref_data.FORM_CODES:
        uid = (request.form.get(f'uid_{code}') or '').strip()
        if uid:
            mapping[code] = uid
            n_soumis += 1
        else:
            mapping.pop(code, None)
    form_mapping.save(mapping)
    flash(f"Correspondance enregistrée pour {n_soumis} formulaire(s) sur {len(ref_data.FORM_CODES)}.", 'success')
    return redirect(url_for('completude'))


@app.route('/completude/calculer', methods=['POST'])
def completude_calculer():
    token, instance = _kobo_credentials()
    mapping = form_mapping.load()
    if not token:
        flash("Connectez-vous d'abord à KoboToolbox.", 'warning')
        return redirect(url_for('completude'))
    if not mapping:
        flash("Associez au moins un formulaire avant de calculer.", 'warning')
        return redirect(url_for('completude'))

    # Filet de sécurité : un même formulaire Kobo mappé sur 2 codes SPAD
    # différents est toujours une erreur de sélection (jamais un cas
    # légitime) — elle fausserait silencieusement les taux des deux
    # formulaires concernés si on la laissait passer. Voir
    # modules/reference_data.py::validate_mapping().
    assets_result = list_assets(token, instance=instance)
    assets = assets_result.get('assets', []) if assets_result.get('success') else []
    mapping_erreurs, _ = ref_data.validate_mapping(mapping, assets)
    if mapping_erreurs:
        for e in mapping_erreurs:
            flash(f"Correspondance incorrecte : {e}", 'danger')
        return redirect(url_for('completude'))

    ref = ref_data.load()
    form_dataframes = {}
    errors = []
    for code, uid in mapping.items():
        res = kobo_load_data(token, uid, instance=instance)
        if res.get('success'):
            form_dataframes[code] = res['df']
        else:
            errors.append(f"{code} : {res.get('error', 'erreur inconnue')}")

    # Calcul en une seule passe sur les données tirées — évite de retirer les
    # mêmes formulaires depuis Kobo pour chaque vue (nationale/région/district).
    # Résultat écrit sur disque (voir _load_completude_cache) : bien trop
    # volumineux pour un cookie de session.
    national = cp.national_summary(ref, form_dataframes)
    cache = {
        'national':       national,
        'district':       cp.district_table(ref, form_dataframes),
        'region':         cp.region_table(ref, form_dataframes),
        'etablissement':  cp.etablissement_table(ref, form_dataframes),
        'enqueteur':      cp.enqueteur_table(ref, form_dataframes),
        'superviseur':    cp.superviseur_table(ref, form_dataframes),
        'anomalies_zero':   cp.all_anomalies_zero(ref, form_dataframes),
        'anomalies_excess': cp.all_anomalies_excess(ref, form_dataframes),
        'export':           cp.export_rows(ref, form_dataframes),
    }
    tendance.add_snapshot(national)  # un point d'historique par jour calendaire (voir modules/tendance.py)
    old_path = session.get('completude_path')
    if old_path and os.path.exists(old_path):
        try: os.remove(old_path)
        except Exception: pass
    fname = f"completude_{uuid.uuid4().hex[:8]}.json"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    with open(save_path, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False)
    session['completude_path'] = save_path
    session['completude_computed_at'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    if errors:
        flash("Calcul partiel — erreurs : " + " · ".join(errors), 'warning')
    else:
        flash(f"Complétude calculée pour {len(form_dataframes)} formulaire(s).", 'success')
    return redirect(url_for('completude'))


# ─── Actualisation automatique KoboToolbox (polling en tâche de fond) ─────────
#
# Application desktop mono-utilisateur, sans URL publique : impossible de
# recevoir un webhook KoboToolbox en conditions réelles (voir note sur
# /webhook/kobo plus bas). L'actualisation « quasi temps réel » repose donc
# sur un sondage périodique de l'API, dans un thread dédié (modules/kobo_sync.py).
# Les nouvelles données ne remplacent jamais silencieusement le jeu de données
# en cours d'analyse : l'utilisateur les applique explicitement.

@app.route('/kobo/sync/start', methods=['POST'])
def kobo_sync_start():
    token    = session.get('kobo_token')
    instance = session.get('kobo_instance')
    uid      = session.get('kobo_uid')
    name     = session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    if not token or not uid:
        return jsonify({"success": False, "error": "Aucun formulaire KoboToolbox actif en session."}), 400

    payload = request.get_json(silent=True) or {}
    try:
        interval = int(payload.get('interval_seconds', 300))
    except (TypeError, ValueError):
        interval = 300

    baseline = (session.get('data_meta') or {}).get('n_obs')
    kobo_sync.start(token, uid, instance, name, interval, baseline)
    return jsonify({"success": True, "status": kobo_sync.status()})


@app.route('/kobo/sync/stop', methods=['POST'])
def kobo_sync_stop():
    kobo_sync.stop()
    return jsonify({"success": True})


@app.route('/kobo/sync/status')
def kobo_sync_status():
    return jsonify(kobo_sync.status())


@app.route('/kobo/sync/apply', methods=['POST'])
def kobo_sync_apply():
    """Applique les nouvelles soumissions détectées par le polling en cours."""
    df = kobo_sync.pop_pending_df()
    if df is None:
        return jsonify({"success": False, "error": "Aucune nouvelle donnée à appliquer."}), 400

    name     = session.get('kobo_asset_name', 'Formulaire KoboToolbox')
    old_path = session.get('data_path')
    if old_path and os.path.exists(old_path):
        try: os.remove(old_path)
        except Exception: pass

    fname     = f"kobo_{uuid.uuid4().hex[:8]}.xlsx"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    df.to_excel(save_path, index=False, engine='openpyxl')

    session['data_path']           = save_path
    session['data_meta']           = summarize_dataframe(df)
    session['data_meta']['source'] = 'KoboToolbox'
    session['data_meta']['name']   = name
    session.pop('child_path', None)

    return jsonify({"success": True, "n_obs": len(df), "n_vars": int(df.shape[1])})


# ─── Webhook KoboToolbox (REST Services) ──────────────────────────────────────
#
# NOTE : cet endpoint ne peut recevoir de requêtes que si l'application est
# exposée sur une URL publique (ex. tunnel ngrok/Cloudflare) — non configuré
# à ce jour. En l'état, KoboToolbox (hébergé dans le cloud) ne peut pas
# atteindre 127.0.0.1 sur le poste de l'utilisateur : ce chemin est inactif
# tant qu'aucun tunnel n'est mis en place. Le mécanisme actif est le polling
# ci-dessus (/kobo/sync/*). À corriger avant toute mise en service de ce
# webhook : il écrit dans `session`, qui est liée au cookie de la requête
# entrante (celle de KoboToolbox) et non à la session du navigateur de
# l'utilisateur — la mise à jour ne serait donc pas visible dans l'interface.

@app.route('/webhook/kobo', methods=['GET', 'POST'])
def kobo_webhook():
    """
    Endpoint reçu par KoboToolbox REST Services à chaque nouvelle soumission.
    GET  → retourne 200 + JSON de confirmation (test de connectivité KoboToolbox).
    POST → reçoit la soumission JSON, la sauvegarde et recharge le DataFrame en session.
    """
    if request.method == 'GET':
        return jsonify({'status': 'ok', 'service': 'SPAD Analyzer Webhook'}), 200

    payload = request.get_json(silent=True) or request.form.to_dict()
    if not payload:
        return jsonify({'status': 'error', 'message': 'Corps vide'}), 400

    # Accumulation des soumissions dans un fichier NDJSON local
    webhook_dir  = os.path.join(app.config['UPLOAD_FOLDER'], 'webhook')
    os.makedirs(webhook_dir, exist_ok=True)
    webhook_file = os.path.join(webhook_dir, 'submissions.jsonl')

    with open(webhook_file, 'a', encoding='utf-8') as f:
        f.write(json.dumps(payload, ensure_ascii=False) + '\n')

    # Reconstruction du DataFrame et mise à jour de la session
    try:
        rows = []
        with open(webhook_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
        if rows:
            df_raw   = pd.DataFrame(rows)
            df_clean = clean_kobo_dataframe(df_raw)
            fname    = f"webhook_{uuid.uuid4().hex[:8]}.xlsx"
            fpath    = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            df_clean.to_excel(fpath, index=False)
            meta = summarize_dataframe(df_clean)
            session['data_path']          = fpath
            session['original_filename']  = 'KoboToolbox (webhook)'
            session['data_meta']          = meta
    except Exception as e:
        app.logger.error(f'Webhook rebuild error: {e}')

    return jsonify({'status': 'ok', 'received': True}), 200


# ─── API JSON (AJAX) ──────────────────────────────────────────────────────────

@app.route('/api/kobo/assets')
def api_kobo_assets():
    """Endpoint JSON — liste les formulaires (appel AJAX)."""
    token = session.get('kobo_token')
    if not token:
        return jsonify({'success': False, 'error': 'Non connecté.'})
    return jsonify(list_assets(token))


@app.route('/api/kobo/asset_info')
def api_kobo_asset_info():
    """Endpoint JSON — info d'un formulaire par UID (appel AJAX)."""
    token = session.get('kobo_token')
    uid   = request.args.get('uid', '')
    if not token or not uid:
        return jsonify({'success': False, 'error': 'Paramètres manquants.'})
    return jsonify(get_asset_info(token, uid))


@app.route('/kobo/diagnostic', methods=['GET', 'POST'])
def kobo_diagnostic():
    """Diagnostic KoboToolbox — teste chaque instance × endpoint."""
    from modules.kobo_connector import diagnose as kobo_diagnose
    results = []
    token   = ''
    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        if token:
            results = kobo_diagnose(token)

    rows_html = ''
    global_ok = False

    for inst in results:
        base      = inst.get('instance', '')
        reachable = inst.get('reachable', False)
        username  = inst.get('username', '')
        label     = '🌍 ONU/WHO/GPEI' if 'humanitarian' in base else '🌐 Publique'

        ep_rows = ''
        for ep in inst.get('endpoints', []):
            s  = ep.get('status', '')
            hc = ep.get('http_code', '—')
            er = ep.get('error', '')
            if s == 'ok':
                icon = '✅'; color = '#17A589'; global_ok = True
            elif 'invalide' in s:
                icon = '🔑'; color = '#E67E22'
            elif s == 'injoignable':
                icon = '🚫'; color = '#C0392B'
            elif s == 'ssl_error':
                icon = '🔒'; color = '#E67E22'
            elif s == 'timeout':
                icon = '⏱'; color = '#E67E22'
            else:
                icon = '❓'; color = '#85929E'

            ep_rows += f"""
            <tr>
              <td style="font-family:monospace;font-size:.78rem">{ep['endpoint']}</td>
              <td><span style="color:{color};font-weight:700">{icon} {s}</span></td>
              <td class="text-center"><span class="badge bg-secondary">{hc}</span></td>
              <td style="font-size:.75rem;color:#C0392B">{er}</td>
            </tr>"""

        border = '#17A589' if reachable else '#C0392B'
        rows_html += f"""
        <div class="card mb-3" style="border-left:4px solid {border}">
          <div class="card-header d-flex justify-content-between align-items-center"
               style="background:#F8F9FA">
            <span style="font-weight:700;color:#1A5276">{label} &nbsp;·&nbsp; {base}</span>
            {'<span style="color:#17A589;font-weight:700">✅ Connecté — ' + username + '</span>'
              if username else
              '<span style="color:#C0392B">✗ Non connecté</span>'}
          </div>
          <div class="card-body p-0">
            <table class="table table-sm mb-0">
              <thead><tr>
                <th>Endpoint testé</th><th>Statut</th>
                <th class="text-center">HTTP</th><th>Détail</th>
              </tr></thead>
              <tbody>{ep_rows}</tbody>
            </table>
          </div>
        </div>"""

    # Conseil final
    conseil = ''
    if results:
        if global_ok:
            conseil = '''<div class="alert alert-success mt-3">
              ✅ <strong>Connexion opérationnelle.</strong>
              Retournez à <a href="/kobo/connect">la page de connexion</a>
              et entrez votre token — il sera accepté.
            </div>'''
        else:
            all_ep_statuses = [
                ep.get('status','')
                for inst in results
                for ep in inst.get('endpoints', [])
            ]
            if any('invalide' in s for s in all_ep_statuses):
                conseil = '''<div class="alert alert-warning mt-3">
                  🔑 <strong>Token invalide ou expiré.</strong><br>
                  Régénérez votre clé API : KoboToolbox → Profil →
                  Paramètres du compte → Sécurité → <strong>Clé API → Régénérer</strong>.
                </div>'''
            elif any(s == 'injoignable' for s in all_ep_statuses):
                conseil = '''<div class="alert alert-danger mt-3">
                  🚫 <strong>Instances injoignables.</strong><br>
                  Vérifiez votre connexion internet. Si vous êtes derrière un
                  proxy ONU/WHO, contactez votre administrateur réseau.
                </div>'''
            else:
                conseil = '''<div class="alert alert-warning mt-3">
                  ❓ <strong>Résultat inattendu.</strong>
                  Copiez les détails ci-dessus et partagez-les pour analyse.
                </div>'''

    html = f"""<!DOCTYPE html>
<html lang="fr"><head>
  <meta charset="UTF-8">
  <title>Diagnostic KoboToolbox</title>
  <link rel="stylesheet"
        href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">
  <style>
    body {{ background:#F0F4F8; font-family:'Inter',sans-serif; padding:30px }}
    .card {{ border-radius:10px; box-shadow:0 1px 6px rgba(0,0,0,.08) }}
    thead th {{ background:#1A5276; color:#fff; font-size:.75rem;
               text-transform:uppercase; letter-spacing:.05em }}
  </style>
</head><body>
<div style="max-width:800px;margin:0 auto">
  <h4 style="color:#1A5276;margin-bottom:4px">🔍 Diagnostic KoboToolbox</h4>
  <p class="text-muted small mb-4">
    Teste chaque instance × endpoint pour identifier la cause exacte.
  </p>

  <div class="card p-4 mb-4">
    <form method="POST">
      <label class="form-label fw-bold">Token API</label>
      <input type="password" name="token" class="form-control mb-3"
             placeholder="Collez votre token ici…" required
             style="font-family:monospace;font-size:.85rem"/>
      <button type="submit" class="btn btn-primary me-2">▶ Lancer le diagnostic</button>
      <a href="/" class="btn btn-outline-secondary">← Application</a>
    </form>
  </div>

  {rows_html}
  {conseil}
</div></body></html>"""

    return html


# ─── XLSForm IMPORT (V2 — handoff depuis PHAKTS·STUDIO) ──────────────────────
# Permet à PHAKTS de pousser son XLSForm directement dans SPAD pour
# prévisualisation de la structure (sans passer par un upload manuel).

@app.after_request
def _add_cors_for_local(resp):
    """Autorise les appels iframe-to-other-port en mode embarqué Electron."""
    origin = request.headers.get('Origin', '')
    if origin.startswith('http://127.0.0.1:') or origin.startswith('http://localhost:'):
        resp.headers['Access-Control-Allow-Origin'] = origin
        resp.headers['Access-Control-Allow-Methods'] = 'POST, GET, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        resp.headers['Access-Control-Allow-Credentials'] = 'true'
        resp.headers['Vary'] = 'Origin'
    return resp


@app.route('/api/xlsform-import', methods=['POST', 'OPTIONS'])
def xlsform_import():
    """Reçoit un XLSForm .xlsx envoyé par PHAKTS, le stocke et le résume."""
    if request.method == 'OPTIONS':
        return ('', 204)

    if 'file' not in request.files:
        return jsonify({'error': "Aucun fichier 'file' dans la requête."}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Nom de fichier vide.'}), 400

    fname = secure_filename(f.filename) or 'phakts_xlsform.xlsx'
    if not fname.lower().endswith('.xlsx'):
        fname += '.xlsx'

    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'xlsforms')
    os.makedirs(target_dir, exist_ok=True)
    saved_path = os.path.join(target_dir, fname)
    f.save(saved_path)

    # Parse the XLSForm pour le récap
    try:
        survey = pd.read_excel(saved_path, sheet_name='survey')
        try:
            choices = pd.read_excel(saved_path, sheet_name='choices')
        except Exception:
            choices = pd.DataFrame()
        try:
            settings = pd.read_excel(saved_path, sheet_name='settings')
        except Exception:
            settings = pd.DataFrame()
    except Exception as e:
        return jsonify({'error': f'XLSForm invalide : {e}'}), 400

    form_title = ''
    if not settings.empty and 'form_title' in settings.columns:
        form_title = str(settings.iloc[0].get('form_title', '') or '')

    session['xlsform_path'] = saved_path
    session['xlsform_title'] = form_title or fname

    return jsonify({
        'ok': True,
        'form_title': form_title or fname,
        'questions': int(len(survey)),
        'choices': int(len(choices)),
        'preview_url': url_for('xlsform_preview', _external=False)
    })


@app.route('/api/kobo/deploy-xlsform', methods=['POST'])
def kobo_deploy_xlsform():
    """Déploie le XLSForm importé depuis PHAKTS dans le compte KoboToolbox de l'utilisateur."""
    saved_path = session.get('xlsform_path')
    if not saved_path or not os.path.exists(saved_path):
        return jsonify({'success': False, 'error': "Aucun XLSForm à déployer. Importez-le d'abord depuis l'onglet Conception PHAKTS."}), 400

    token = session.get('kobo_token')
    if not token:
        return jsonify({'success': False, 'error': "Connectez-vous d'abord à KoboToolbox (onglet Mes formulaires).",
                        'need_kobo_connect': True}), 401

    name = session.get('xlsform_title') or 'PHAKTS XLSForm'
    instance = session.get('kobo_instance')
    custom = session.get('kobo_custom_instance')

    res = deploy_xlsform(token, saved_path, name=name, instance=instance, custom_instance=custom)
    status = 200 if res.get('success') else 502
    return jsonify(res), status


@app.route('/xlsform/preview')
def xlsform_preview():
    """Affiche la structure du XLSForm importé depuis PHAKTS."""
    saved_path = session.get('xlsform_path')
    if not saved_path or not os.path.exists(saved_path):
        return render_template('xlsform_preview.html',
                               imported=False,
                               form_title='', survey=[], choices=[])

    survey_df = pd.read_excel(saved_path, sheet_name='survey').fillna('')
    try:
        choices_df = pd.read_excel(saved_path, sheet_name='choices').fillna('')
    except Exception:
        choices_df = pd.DataFrame()

    return render_template(
        'xlsform_preview.html',
        imported=True,
        form_title=session.get('xlsform_title', ''),
        survey=survey_df.to_dict(orient='records'),
        choices=choices_df.to_dict(orient='records'),
        survey_cols=list(survey_df.columns),
        choices_cols=list(choices_df.columns) if not choices_df.empty else [],
    )


if __name__ == '__main__':
    app.run(debug=True, port=5050)
